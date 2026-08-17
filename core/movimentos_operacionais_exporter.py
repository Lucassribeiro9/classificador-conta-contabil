from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.utils import get_column_letter

from core.movimentos_operacionais_snapshot import (
    LoteOperacionalSnapshot,
    MovimentoOperacionalSnapshotItem,
)


class MovimentoOperacionalExportError(ValueError):
    """Erro de dominio ao gerar planilha classificada de movimentos."""


_ORIGINAL_COLUMNS_BY_LAYOUT: dict[str, tuple[str, ...]] = {
    "operacional_valor_saldo_v1": (
        "data",
        "conta_financeira",
        "historico",
        "valor",
        "saldo",
        "contrapartida",
        "tipo_movimento",
        "documento",
        "observacao",
    ),
    "operacional_debito_credito_saldo_v1": (
        "data",
        "conta_financeira",
        "historico",
        "debito",
        "credito",
        "saldo",
        "contrapartida",
        "tipo_movimento",
        "documento",
        "observacao",
    ),
    "operacional_valor_legado_v1": (
        "data",
        "conta_financeira",
        "historico",
        "valor",
        "contrapartida",
        "tipo_movimento",
        "documento",
        "observacao",
    ),
}

_CONTROL_COLUMNS = (
    "lote_id",
    "movimento_id",
    "linha_original",
    "layout_version",
    "export_revision",
    "row_version",
)

_READ_ONLY_COLUMNS = (
    "contrapartida_sugerida",
    "confidence_sugerida",
    "status_atual",
    "mensagem_validacao",
    "saldo_observado_original",
    "saldo_observado_decimal",
    "saldo_calculado_decimal",
    "warnings_saldo",
)

_EDITABLE_COLUMNS = (
    "decisao_revisao",
    "contrapartida_final",
    "observacao_revisao",
)


def gerar_planilha_classificada(snapshot: LoteOperacionalSnapshot) -> bytes:
    """Gera XLSX classificado a partir de um snapshot operacional."""

    original_columns = _ORIGINAL_COLUMNS_BY_LAYOUT.get(snapshot.layout_version)
    if original_columns is None:
        raise MovimentoOperacionalExportError(
            f"Layout operacional desconhecido: {snapshot.layout_version}"
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Movimentos"

    headers = (
        *original_columns,
        *_CONTROL_COLUMNS,
        *_READ_ONLY_COLUMNS,
        *_EDITABLE_COLUMNS,
    )
    sheet.append(headers)
    for movimento in snapshot.movimentos:
        sheet.append([_value_for_column(movimento, column) for column in headers])

    _format_sheet(sheet, editable_columns=set(_EDITABLE_COLUMNS))

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _value_for_column(
    movimento: MovimentoOperacionalSnapshotItem,
    column: str,
) -> Any:
    """Resolve valor de uma coluna exportavel."""

    if column == "valor":
        return movimento.valor_original
    if column == "debito":
        return movimento.valor_absoluto if movimento.direcao == "credito" else None
    if column == "credito":
        return movimento.valor_absoluto if movimento.direcao == "debito" else None
    if column == "saldo":
        return movimento.saldo_observado_original
    if column == "contrapartida":
        return movimento.contrapartida
    if column == "mensagem_validacao":
        return "; ".join(movimento.mensagem_validacao)
    if column == "warnings_saldo":
        return "; ".join(movimento.warnings_saldo)
    if column in _EDITABLE_COLUMNS:
        if column == "contrapartida_final":
            return movimento.contrapartida_final
        return None
    return getattr(movimento, column)


def _format_sheet(sheet, *, editable_columns: set[str]) -> None:
    """Aplica formatacao operacional minima e protecao visual."""

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.protection = Protection(locked=True)

    headers = [cell.value for cell in sheet[1]]
    for column_index, header in enumerate(headers, start=1):
        letter = get_column_letter(column_index)
        sheet.column_dimensions[letter].width = max(14, min(40, len(str(header)) + 2))
        locked = header not in editable_columns
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row, column=column_index).protection = Protection(
                locked=locked
            )

    sheet.freeze_panes = "A2"
    sheet.protection.sheet = True

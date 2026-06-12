from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class PlanoContasParseError(ValueError):
    """Erro de validacao do arquivo de plano de contas."""


_REQUIRED_COLUMNS = {
    "codigo": "codigo",
    "tipo": "tipo",
    "classificacao": "classificacao",
    "nome": "nome",
    "grau": "grau",
}


def parse_plano_contas_xlsx(path: str | Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active

        header_by_column = _find_header_by_column(sheet.iter_rows(values_only=True))
        contas: list[dict[str, Any]] = []

        for row_number, row in enumerate(
            sheet.iter_rows(
                min_row=header_by_column["row_number"] + 1,
                values_only=True,
            ),
            start=header_by_column["row_number"] + 1,
        ):
            if _is_empty_row(row):
                continue
            contas.append(_parse_account_row(row_number, row, header_by_column))

        return contas
    finally:
        workbook.close()


def _find_header_by_column(rows) -> dict[str, int]:
    for row_number, row in enumerate(rows, start=1):
        normalized_cells = {
            _normalize_header(value): index
            for index, value in enumerate(row)
            if _normalize_header(value)
        }
        if all(column in normalized_cells for column in _REQUIRED_COLUMNS):
            return {
                "row_number": row_number,
                **{
                    field_name: normalized_cells[column_name]
                    for column_name, field_name in _REQUIRED_COLUMNS.items()
                },
            }

    raise PlanoContasParseError(
        "Cabecalho do plano de contas nao encontrado: "
        "esperado Codigo, Tipo, Classificacao, Nome e Grau."
    )


def _parse_account_row(
    row_number: int,
    row: tuple[Any, ...],
    header_by_column: dict[str, int],
) -> dict[str, Any]:
    raw_account = {
        field: _cell(row, column_index)
        for field, column_index in header_by_column.items()
        if field != "row_number"
    }
    missing_fields = [
        field for field, value in raw_account.items() if _is_blank_value(value)
    ]
    if missing_fields:
        raise PlanoContasParseError(
            f"Linha {row_number} incompleta: campos ausentes "
            f"{', '.join(sorted(missing_fields))}."
        )

    tipo = str(raw_account["tipo"]).strip().upper()
    if tipo not in {"A", "S"}:
        raise PlanoContasParseError(
            f"Linha {row_number} invalida: tipo deve ser A ou S."
        )

    try:
        codigo = int(raw_account["codigo"])
        grau = int(raw_account["grau"])
    except (TypeError, ValueError) as exc:
        raise PlanoContasParseError(
            f"Linha {row_number} invalida: codigo e grau devem ser numericos."
        ) from exc

    return {
        "codigo": codigo,
        "tipo": tipo,
        "classificacao": str(raw_account["classificacao"]).strip(),
        "nome": str(raw_account["nome"]).strip(),
        "grau": grau,
    }


def _cell(row: tuple[Any, ...], index: int) -> Any:
    if index >= len(row):
        return None
    return row[index]


def _is_empty_row(row: tuple[Any, ...]) -> bool:
    return all(_is_blank_value(value) for value in row)


def _is_blank_value(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return (
        str(value)
        .strip()
        .lower()
        .replace("ç", "c")
        .replace("ã", "a")
        .replace("á", "a")
        .replace("â", "a")
        .replace("é", "e")
        .replace("ê", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ô", "o")
        .replace("ú", "u")
    )

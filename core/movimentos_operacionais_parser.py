from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook


class MovimentoOperacionalParseError(ValueError):
    """Erro de validacao do layout de movimentos operacionais."""


@dataclass(frozen=True)
class MovimentoOperacionalMetadata:
    empresa_nome: str
    codigo_dominio: str
    cnpj_cpf: str
    periodo_inicio: str
    periodo_fim: str


@dataclass(frozen=True)
class MovimentoOperacionalParseResult:
    metadata: MovimentoOperacionalMetadata
    movimentos: list[dict[str, Any]]
    layout_version: str


_SHEET_NAME = "Movimentos"
_BASE_REQUIRED_COLUMNS = ("data", "conta_financeira", "historico")
_VALUE_COLUMN = "valor"
_SALDO_COLUMN = "saldo"
_DEBITO_COLUMN = "debito"
_CREDITO_COLUMN = "credito"
_OPTIONAL_COLUMNS = (
    "contrapartida",
    "tipo_movimento",
    "documento",
    "observacao",
)
_SYSTEM_COLUMNS = (
    "status_sugerido",
    "confidence_sugerida",
    "mensagem_validacao",
)
_LAYOUT_COLUMNS = (
    *_BASE_REQUIRED_COLUMNS,
    _VALUE_COLUMN,
    _SALDO_COLUMN,
    _DEBITO_COLUMN,
    _CREDITO_COLUMN,
)
_COLUMN_ALIASES = {
    column: column
    for column in (*_LAYOUT_COLUMNS, *_OPTIONAL_COLUMNS, *_SYSTEM_COLUMNS)
}


def parse_movimentos_operacionais_xlsx(
    path: str | Path,
) -> MovimentoOperacionalParseResult:
    """Le planilha operacional .xlsx e retorna metadados e movimentos."""

    file_path = Path(path)
    if file_path.suffix.lower() != ".xlsx":
        raise MovimentoOperacionalParseError(
            "Arquivo de movimentos operacionais deve estar no formato .xlsx."
        )

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        if _SHEET_NAME not in workbook.sheetnames:
            raise MovimentoOperacionalParseError(
                "Layout de movimentos operacionais invalido: aba Movimentos "
                "nao encontrada."
            )

        sheet = workbook[_SHEET_NAME]
        metadata_values: dict[str, Any] = {}
        header_by_column: dict[str, int] | None = None
        layout_version: str | None = None
        movimentos: list[dict[str, Any]] = []

        for row in sheet.iter_rows(values_only=True):
            if _is_empty_row(row):
                continue

            metadata_values.update(_extract_metadata(row))

            maybe_header = _header_by_column(row)
            if maybe_header is not None:
                header_by_column = maybe_header
                layout_version = _detect_layout_version(header_by_column)
                continue

            if header_by_column is None:
                continue

            movimento = _parse_movimento_row(row, header_by_column)
            if movimento is not None:
                movimentos.append(movimento)

        metadata = _build_metadata(metadata_values)
        if header_by_column is None:
            raise MovimentoOperacionalParseError(
                "Layout de movimentos operacionais invalido: cabecalho com "
                "colunas obrigatorias nao encontrado."
            )

        assert layout_version is not None
        return MovimentoOperacionalParseResult(
            metadata=metadata,
            movimentos=movimentos,
            layout_version=layout_version,
        )
    finally:
        workbook.close()


def _extract_metadata(row: tuple[Any, ...]) -> dict[str, Any]:
    """Extrai metadados reconhecidos em uma linha por rotulo."""

    metadata: dict[str, Any] = {}
    for index, value in enumerate(row):
        label = _normalize_text(value)
        if label == "empresa":
            metadata["empresa_nome"] = _first_clean_text_after(row, index)
        elif label == "codigo dominio":
            metadata["codigo_dominio"] = _first_clean_text_after(row, index)
        elif label == "cnpj/cpf":
            metadata["cnpj_cpf"] = _only_digits(_first_clean_text_after(row, index))
        elif label == "periodo inicio":
            metadata["periodo_inicio"] = _parse_br_date(
                _first_clean_text_after(row, index)
            )
        elif label == "periodo fim":
            metadata["periodo_fim"] = _parse_br_date(
                _first_clean_text_after(row, index)
            )
    return metadata


def _first_clean_text_after(row: tuple[Any, ...], index: int) -> str | None:
    """Retorna o primeiro texto preenchido apos a coluna informada."""

    for value in row[index + 1 :]:
        text = _clean_text(value)
        if text:
            return text
    return None


def _header_by_column(row: tuple[Any, ...]) -> dict[str, int] | None:
    """Detecta o cabecalho por colunas base e decisorias de layout."""

    header_by_field: dict[str, int] = {}
    for index, value in enumerate(row):
        field_name = _COLUMN_ALIASES.get(_normalize_text(value))
        if field_name is not None:
            header_by_field[field_name] = index

    if not all(column in header_by_field for column in _BASE_REQUIRED_COLUMNS):
        return None

    if _detectable_layout(header_by_field):
        return header_by_field

    return None


def _detectable_layout(header_by_column: dict[str, int]) -> bool:
    """Indica se o cabecalho contem uma assinatura de layout conhecida."""

    has_valor = _VALUE_COLUMN in header_by_column
    has_saldo = _SALDO_COLUMN in header_by_column
    has_debito = _DEBITO_COLUMN in header_by_column
    has_credito = _CREDITO_COLUMN in header_by_column

    if has_valor and (has_debito or has_credito):
        return True
    if has_valor:
        return True
    return has_debito and has_credito and has_saldo


def _detect_layout_version(header_by_column: dict[str, int]) -> str:
    """Classifica a versao conceitual do layout pelo cabecalho."""

    has_valor = _VALUE_COLUMN in header_by_column
    has_saldo = _SALDO_COLUMN in header_by_column
    has_debito = _DEBITO_COLUMN in header_by_column
    has_credito = _CREDITO_COLUMN in header_by_column

    if has_valor and (has_debito or has_credito):
        raise MovimentoOperacionalParseError(
            "Layout de movimentos operacionais ambiguo: cabecalho contem valor "
            "junto com debito ou credito."
        )
    if has_valor and has_saldo:
        return "operacional_valor_saldo_v1"
    if has_valor:
        return "operacional_valor_legado_v1"
    if has_debito and has_credito and has_saldo:
        return "operacional_debito_credito_saldo_v1"

    raise MovimentoOperacionalParseError(
        "Layout de movimentos operacionais invalido: cabecalho com colunas "
        "obrigatorias nao encontrado."
    )


def _parse_movimento_row(
    row: tuple[Any, ...],
    header_by_column: dict[str, int],
) -> dict[str, Any] | None:
    """Converte uma linha de planilha em movimento operacional bruto."""

    raw = {
        field: _cell(row, header_by_column.get(field))
        for field in (
            *_BASE_REQUIRED_COLUMNS,
            _VALUE_COLUMN,
            _DEBITO_COLUMN,
            _CREDITO_COLUMN,
            *_OPTIONAL_COLUMNS,
        )
    }
    if all(_is_blank_value(value) for value in raw.values()):
        return None

    data = _format_date(raw["data"])
    if data is None:
        return None

    return {
        "data": data,
        "conta_financeira": _clean_integer_like_number(raw["conta_financeira"]),
        "historico": _clean_text(raw["historico"]),
        "valor": raw["valor"],
        "contrapartida": _clean_integer_like_number(raw["contrapartida"]),
        "tipo_movimento": _clean_text(raw["tipo_movimento"]),
        "documento": _clean_text(raw["documento"]),
        "observacao": _clean_text(raw["observacao"]),
    }


def _build_metadata(metadata_values: dict[str, Any]) -> MovimentoOperacionalMetadata:
    """Valida metadados obrigatorios e monta o contrato de retorno."""

    missing_fields = [
        field
        for field in ("cnpj_cpf", "periodo_inicio", "periodo_fim")
        if _is_blank_value(metadata_values.get(field))
    ]
    if missing_fields:
        raise MovimentoOperacionalParseError(
            "Lote de movimentos operacionais sem metadados obrigatorios: "
            + ", ".join(missing_fields)
            + "."
        )

    return MovimentoOperacionalMetadata(
        empresa_nome=_clean_text(metadata_values.get("empresa_nome")) or "",
        codigo_dominio=_clean_text(metadata_values.get("codigo_dominio")) or "",
        cnpj_cpf=str(metadata_values["cnpj_cpf"]),
        periodo_inicio=str(metadata_values["periodo_inicio"]),
        periodo_fim=str(metadata_values["periodo_fim"]),
    )


def _cell(row: tuple[Any, ...], index: int | None) -> Any:
    """Busca uma celula com tolerancia a indice ausente ou fora da linha."""

    if index is None or index >= len(row):
        return None
    return row[index]


def _format_date(value: Any) -> str | None:
    """Normaliza datas de celulas Excel ou texto brasileiro para ISO."""

    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return _parse_br_date(value)


def _parse_br_date(value: Any) -> str | None:
    """Converte data em formato dd/mm/aaaa para ISO."""

    text = _clean_text(value)
    if not text:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    try:
        day, month, year = text.split("/")
        parsed_date = date(int(year), int(month), int(day))
    except ValueError:
        return None
    return parsed_date.isoformat()


def _only_digits(value: Any) -> str | None:
    """Remove caracteres nao numericos de um valor textual."""

    text = _clean_text(value)
    if text is None:
        return None
    return re.sub(r"\D", "", text)


def _clean_integer_like_number(value: Any) -> Any:
    """Converte codigos numericos de planilha para inteiros quando possivel."""

    text = _clean_text(value)
    if text is None:
        return None
    if re.fullmatch(r"\d+\.0+", text):
        return int(text.split(".", 1)[0])
    if re.fullmatch(r"\d+", text):
        return int(text)
    return value


def _is_empty_row(row: tuple[Any, ...]) -> bool:
    """Indica se uma linha esta completamente vazia."""

    return all(_is_blank_value(value) for value in row)


def _is_blank_value(value: Any) -> bool:
    """Indica se um valor deve ser tratado como vazio."""

    return value is None or str(value).strip() == ""


def _clean_text(value: Any) -> str | None:
    """Retorna texto sem espacos externos ou None para valores vazios."""

    if _is_blank_value(value):
        return None
    return str(value).strip()


def _normalize_text(value: Any) -> str:
    """Normaliza texto para comparacoes simples de rotulos e cabecalhos."""

    if value is None:
        return ""
    return (
        str(value)
        .strip()
        .lower()
        .replace("ç", "c")
        .replace("ã", "a")
        .replace("á", "a")
        .replace("â", "a")
        .replace("é", "e")
        .replace("ê", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ô", "o")
        .replace("ú", "u")
    )

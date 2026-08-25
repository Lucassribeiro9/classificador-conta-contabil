from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from core.audit import record_audit_event
from core.models import MovimentoOperacionalImportado
from core.movimentos_operacionais_review import (
    MovimentoReviewError,
    review_movimento_operacional,
)


class MovimentoOperacionalFeedbackImportError(ValueError):
    """Erro estrutural ao importar feedback por planilha classificada."""


REQUIRED_COLUMNS = {
    "lote_id",
    "movimento_id",
    "linha_original",
    "layout_version",
    "export_revision",
    "row_version",
    "decisao_revisao",
    "contrapartida_final",
    "observacao_revisao",
}
READ_ONLY_COLUMNS = {
    "contrapartida_sugerida",
    "confidence_sugerida",
    "status_atual",
    "mensagem_validacao",
    "saldo_observado_original",
    "saldo_observado_decimal",
    "saldo_calculado_decimal",
    "warnings_saldo",
}
ACTION_MAP = {
    "aprovar": "approve",
    "corrigir": "correct",
    "rejeitar": "reject",
}


@dataclass(frozen=True)
class ResultadoLinhaFeedbackOperacional:
    linha_original: int | None
    movimento_id: int | None
    status: str
    mensagem: str


@dataclass(frozen=True)
class ResumoFeedbackOperacional:
    total_linhas: int
    total_aplicado: int
    total_ignorado: int
    total_invalido: int
    total_conflitante: int
    total_nao_autorizado: int
    resultados: list[ResultadoLinhaFeedbackOperacional]


def importar_feedback_planilha_classificada(
    db: Session,
    path: str | Path,
    *,
    empresa_id: int,
    lote_id: int,
    usuario_id: int | None,
    actor_metadata: dict[str, Any] | None = None,
) -> ResumoFeedbackOperacional:
    """Importa revisoes em lote de planilha classificada com resultado parcial."""
    rows = _read_feedback_rows(path)
    resultados: list[ResultadoLinhaFeedbackOperacional] = []
    expected_export_revision: str | None = None

    for row in rows:
        export_revision = _normalize_text(row.get("export_revision"))
        if (
            expected_export_revision is None
            and export_revision
            and _has_valid_control_values(row)
        ):
            expected_export_revision = export_revision

        resultado = _process_feedback_row(
            db,
            row,
            empresa_id=empresa_id,
            lote_id=lote_id,
            usuario_id=usuario_id,
            actor_metadata=actor_metadata,
            expected_export_revision=expected_export_revision,
        )
        resultados.append(resultado)

    resumo = _build_summary(resultados)
    record_audit_event(
        db,
        event_type="operational_movements.feedback_imported",
        user_id=usuario_id,
        empresa_id=empresa_id,
        resource_id=str(lote_id),
        metadata={
            **(actor_metadata or {"actor_type": "user"}),
            "lote_id": lote_id,
            "total_linhas": resumo.total_linhas,
            "total_aplicado": resumo.total_aplicado,
            "total_ignorado": resumo.total_ignorado,
            "total_invalido": resumo.total_invalido,
            "total_conflitante": resumo.total_conflitante,
            "total_nao_autorizado": resumo.total_nao_autorizado,
        },
    )
    return resumo


def _read_feedback_rows(path: str | Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    try:
        if "Movimentos" not in workbook.sheetnames:
            raise MovimentoOperacionalFeedbackImportError(
                "Aba Movimentos não encontrada"
            )
        sheet = workbook["Movimentos"]
        headers = [cell.value for cell in sheet[1]]
        if not headers or any(header is None for header in headers):
            raise MovimentoOperacionalFeedbackImportError(
                "Cabeçalho da planilha inválido"
            )

        header_names = [str(header) for header in headers]
        missing = sorted(REQUIRED_COLUMNS - set(header_names))
        if missing:
            raise MovimentoOperacionalFeedbackImportError(
                "Colunas obrigatórias ausentes: " + ", ".join(missing)
            )

        rows: list[dict[str, Any]] = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            if all(value is None for value in values):
                continue
            rows.append(dict(zip(header_names, values)))
        return rows
    finally:
        workbook.close()


def _process_feedback_row(
    db: Session,
    row: dict[str, Any],
    *,
    empresa_id: int,
    lote_id: int,
    usuario_id: int | None,
    actor_metadata: dict[str, Any] | None,
    expected_export_revision: str | None,
) -> ResultadoLinhaFeedbackOperacional:
    linha_original = _optional_int(row.get("linha_original"))
    movimento_id = _optional_int(row.get("movimento_id"))

    if not _has_valid_control_values(row):
        return ResultadoLinhaFeedbackOperacional(
            linha_original=linha_original,
            movimento_id=movimento_id,
            status="invalida",
            mensagem="Campos de controle inválidos",
        )

    if (
        expected_export_revision is not None
        and _normalize_text(row.get("export_revision")) != expected_export_revision
    ):
        return ResultadoLinhaFeedbackOperacional(
            linha_original=linha_original,
            movimento_id=movimento_id,
            status="invalida",
            mensagem="export_revision divergente no arquivo",
        )

    if int(row["lote_id"]) != lote_id:
        return _unauthorized_row(linha_original, movimento_id)

    movimento = db.get(MovimentoOperacionalImportado, movimento_id)
    if (
        movimento is None
        or movimento.empresa_id != empresa_id
        or movimento.lote_id != lote_id
    ):
        return _unauthorized_row(linha_original, movimento_id)

    decisao = _normalize_text(row.get("decisao_revisao"))
    if not decisao:
        return ResultadoLinhaFeedbackOperacional(
            linha_original=linha_original,
            movimento_id=movimento_id,
            status="ignorada",
            mensagem="Linha sem decisão de revisão",
        )
    if decisao not in ACTION_MAP:
        return ResultadoLinhaFeedbackOperacional(
            linha_original=linha_original,
            movimento_id=movimento_id,
            status="invalida",
            mensagem="Decisão de revisão inválida",
        )

    row_version = int(row["row_version"])
    if row_version != movimento.row_version:
        if _is_idempotent_replay(row, movimento, decisao):
            return ResultadoLinhaFeedbackOperacional(
                linha_original=linha_original,
                movimento_id=movimento_id,
                status="ignorada",
                mensagem="Decisão já aplicada anteriormente",
            )
        return ResultadoLinhaFeedbackOperacional(
            linha_original=linha_original,
            movimento_id=movimento_id,
            status="conflitante",
            mensagem="Linha desatualizada em relação ao movimento atual",
        )

    if _readonly_changed(row, movimento):
        return ResultadoLinhaFeedbackOperacional(
            linha_original=linha_original,
            movimento_id=movimento_id,
            status="invalida",
            mensagem="Campo somente leitura alterado",
        )

    try:
        review_movimento_operacional(
            db=db,
            movimento_id=movimento.id,
            empresa_id=empresa_id,
            usuario_id=usuario_id,
            actor_metadata=actor_metadata,
            action=ACTION_MAP[decisao],
            conta_final=_optional_int(row.get("contrapartida_final")),
        )
    except MovimentoReviewError as exc:
        return ResultadoLinhaFeedbackOperacional(
            linha_original=linha_original,
            movimento_id=movimento_id,
            status="invalida",
            mensagem=str(exc),
        )

    return ResultadoLinhaFeedbackOperacional(
        linha_original=linha_original,
        movimento_id=movimento_id,
        status="aplicada",
        mensagem="Decisão aplicada",
    )


def _is_idempotent_replay(
    row: dict[str, Any],
    movimento: MovimentoOperacionalImportado,
    decisao: str,
) -> bool:
    if movimento.row_version != int(row["row_version"]) + 1:
        return False

    conta_final = _optional_int(row.get("contrapartida_final"))
    if decisao == "aprovar":
        return (
            movimento.status == "aprovado"
            and movimento.contrapartida_final == conta_final
        )
    if decisao == "corrigir":
        return (
            movimento.status == "corrigido"
            and movimento.contrapartida_final == conta_final
        )
    if decisao == "rejeitar":
        return movimento.status == "rejeitado" and movimento.contrapartida_final is None
    return False


def _has_valid_control_values(row: dict[str, Any]) -> bool:
    for column in ("lote_id", "movimento_id", "linha_original", "row_version"):
        if _optional_int(row.get(column)) is None:
            return False
    return bool(_normalize_text(row.get("layout_version"))) and bool(
        _normalize_text(row.get("export_revision"))
    )


def _readonly_changed(
    row: dict[str, Any],
    movimento: MovimentoOperacionalImportado,
) -> bool:
    expected_values = {
        "contrapartida_sugerida": movimento.contrapartida_sugerida,
        "confidence_sugerida": movimento.confidence_sugerida,
        "status_atual": movimento.status,
        "mensagem_validacao": "; ".join(movimento.mensagens_validacao or []),
        "saldo_observado_original": movimento.saldo_observado_original,
        "saldo_observado_decimal": movimento.saldo_observado_decimal,
        "saldo_calculado_decimal": movimento.saldo_calculado_decimal,
        "warnings_saldo": "; ".join(movimento.warnings_saldo or []),
    }
    for column in READ_ONLY_COLUMNS:
        if column not in row:
            continue
        value = row.get(column)
        if _is_empty(value):
            continue
        expected = expected_values.get(column)
        if _normalize_compare(value) != _normalize_compare(expected):
            return True
    return False


def _unauthorized_row(
    linha_original: int | None,
    movimento_id: int | None,
) -> ResultadoLinhaFeedbackOperacional:
    return ResultadoLinhaFeedbackOperacional(
        linha_original=linha_original,
        movimento_id=movimento_id,
        status="nao_autorizada",
        mensagem="Linha fora do escopo da empresa/lote",
    )


def _build_summary(
    resultados: list[ResultadoLinhaFeedbackOperacional],
) -> ResumoFeedbackOperacional:
    return ResumoFeedbackOperacional(
        total_linhas=len(resultados),
        total_aplicado=sum(item.status == "aplicada" for item in resultados),
        total_ignorado=sum(item.status == "ignorada" for item in resultados),
        total_invalido=sum(item.status == "invalida" for item in resultados),
        total_conflitante=sum(item.status == "conflitante" for item in resultados),
        total_nao_autorizado=sum(
            item.status == "nao_autorizada" for item in resultados
        ),
        resultados=resultados,
    )


def _optional_int(value: Any) -> int | None:
    if _is_empty(value):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _normalize_text(value: Any) -> str:
    if _is_empty(value):
        return ""
    return str(value).strip().lower()


def _normalize_compare(value: Any) -> str:
    if _is_empty(value):
        return ""
    if isinstance(value, float):
        return f"{value:.10g}"
    return str(value).strip()


def _is_empty(value: Any) -> bool:
    return value is None or str(value).strip() == ""

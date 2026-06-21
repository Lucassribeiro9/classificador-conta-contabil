from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook


class RazaoParseError(ValueError):
    """Erro de validacao do arquivo do livro-razao."""


@dataclass(frozen=True)
class RazaoMetadata:
    empresa_nome: str
    cnpj_cpf: str
    periodo_inicio: str
    periodo_fim: str


@dataclass(frozen=True)
class RazaoParseResult:
    metadata: RazaoMetadata
    lancamentos: list[dict[str, Any]]


_REQUIRED_COLUMNS = ("data", "historico", "contrapartida", "debito", "credito")
_OPTIONAL_COLUMNS = ("numero", "conta_origem")
_COLUMN_ALIASES = {
    "data": "data",
    "numero": "numero",
    "conta_origem": "conta_origem",
    "historico": "historico",
    "contrapartida": "contrapartida",
    "cta.c.part.": "contrapartida",
    "debito": "debito",
    "credito": "credito",
}


def parse_razao_xlsx(path: str | Path) -> list[dict[str, Any]]:
    return _parse_razao_xlsx(path, require_metadata=False).lancamentos


def parse_razao_xlsx_with_metadata(path: str | Path) -> RazaoParseResult:
    return _parse_razao_xlsx(path, require_metadata=True)


def _parse_razao_xlsx(path: str | Path, *, require_metadata: bool) -> RazaoParseResult:
    file_path = Path(path)
    if file_path.suffix.lower() != ".xlsx":
        raise RazaoParseError("Arquivo do razao deve estar no formato .xlsx.")

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        conta_origem: str | None = None
        header_by_column: dict[str, int | None] | None = None
        lancamentos: list[dict[str, Any]] = []
        empresa_nome: str | None = None
        cnpj_cpf: str | None = None
        periodo_inicio: str | None = None
        periodo_fim: str | None = None

        for row in sheet.iter_rows(values_only=True):
            if _is_empty_row(row):
                continue

            row_metadata = _extract_metadata(row)
            if row_metadata.get("empresa_nome") is not None:
                empresa_nome = row_metadata["empresa_nome"]
            if row_metadata.get("cnpj_cpf") is not None:
                cnpj_cpf = row_metadata["cnpj_cpf"]
            if row_metadata.get("periodo_inicio") is not None:
                periodo_inicio = row_metadata["periodo_inicio"]
            if row_metadata.get("periodo_fim") is not None:
                periodo_fim = row_metadata["periodo_fim"]

            conta_bloco = _extract_account_block(row)
            if conta_bloco is not None:
                conta_origem = conta_bloco
                continue

            maybe_header = _header_by_column(row)
            if maybe_header is not None:
                header_by_column = maybe_header
                continue

            if header_by_column is None:
                continue

            if conta_origem is None and header_by_column.get("conta_origem") is None:
                continue

            if _is_balance_row(row):
                continue

            lancamento = _parse_entry_row(row, conta_origem, header_by_column)
            if lancamento is not None:
                lancamentos.append(lancamento)

        metadata = _build_metadata(
            empresa_nome,
            cnpj_cpf,
            periodo_inicio,
            periodo_fim,
            require_metadata=require_metadata,
        )
        if require_metadata and header_by_column is None and not lancamentos:
            raise RazaoParseError(
                "Layout do razao nao reconhecido: cabecalho obrigatorio nao "
                "encontrado. Verifique colunas como data, historico, "
                "contrapartida, debito e credito."
            )
        return RazaoParseResult(metadata=metadata, lancamentos=lancamentos)
    finally:
        workbook.close()


def normalize_lancamento_razao(lancamento: dict[str, Any]) -> dict[str, Any]:
    debito = lancamento.get("debito")
    credito = lancamento.get("credito")
    contrapartida = lancamento.get("contrapartida")
    conta_origem = lancamento.get("conta_origem")

    has_debito = not _is_blank_value(debito)
    has_credito = not _is_blank_value(credito)
    if has_debito == has_credito:
        raise RazaoParseError(
            "Linha do razao deve possuir debito ou credito valido."
        )

    if has_debito:
        conta_debito = conta_origem
        conta_credito = contrapartida
        direcao = "debito"
        valor = debito
    else:
        conta_debito = contrapartida
        conta_credito = conta_origem
        direcao = "credito"
        valor = credito

    return {
        "conta_origem": conta_origem,
        "conta_contrapartida": contrapartida,
        "conta_debito": conta_debito,
        "conta_credito": conta_credito,
        "direcao": direcao,
        "data": lancamento.get("data"),
        "numero": lancamento.get("numero"),
        "historico": lancamento.get("historico"),
        "valor": valor,
    }


def normalize_razao_historico(historico: Any) -> str:
    return " ".join(str(historico).strip().lower().split())


def build_razao_dedup_key(lancamento: dict[str, Any]) -> tuple[Any, ...]:
    historico_normalizado = lancamento.get("historico_normalizado")
    if _is_blank_value(historico_normalizado):
        historico_normalizado = normalize_razao_historico(
            lancamento.get("historico", "")
        )

    return (
        lancamento.get("empresa_id"),
        lancamento.get("numero_lancamento"),
        lancamento.get("data"),
        lancamento.get("conta_origem"),
        lancamento.get("conta_contrapartida"),
        lancamento.get("valor"),
        lancamento.get("direcao"),
        historico_normalizado,
    )


def _extract_account_block(row: tuple[Any, ...]) -> str | None:
    for index, value in enumerate(row):
        text = _clean_text(value)
        if not text:
            continue

        normalized = _normalize_text(text)
        if normalized == "conta:" or normalized == "conta":
            return _first_text_after(row, index)

        if normalized.startswith("conta:"):
            account = text.split(":", 1)[1].strip()
            if account:
                return account.split()[0]

    return None


def _first_text_after(row: tuple[Any, ...], index: int) -> str | None:
    for value in row[index + 1 :]:
        text = _clean_integer_like_text(value)
        if text:
            return text.split()[0]
    return None


def _extract_metadata(row: tuple[Any, ...]) -> dict[str, str | None]:
    metadata: dict[str, str | None] = {}
    for index, value in enumerate(row):
        text = _clean_text(value)
        if not text:
            continue

        normalized = _normalize_text(text)
        if normalized in {"empresa:", "empresa"}:
            metadata["empresa_nome"] = _first_clean_text_after(row, index)
        elif normalized in {"c.n.p.j.:", "c.n.p.j.", "cnpj:", "cnpj"}:
            cnpj = _first_clean_text_after(row, index)
            metadata["cnpj_cpf"] = _only_digits(cnpj)
        elif normalized in {"periodo inicio:", "periodo inicio"}:
            metadata["periodo_inicio"] = _parse_br_date(
                _first_clean_text_after(row, index)
            )
        elif normalized in {"periodo fim:", "periodo fim"}:
            metadata["periodo_fim"] = _parse_br_date(
                _first_clean_text_after(row, index)
            )
        elif normalized in {"periodo:", "periodo"}:
            periodo = _first_clean_text_after(row, index)
            inicio, fim = _parse_period_range(periodo)
            metadata["periodo_inicio"] = inicio
            metadata["periodo_fim"] = fim

    return metadata


def _first_clean_text_after(row: tuple[Any, ...], index: int) -> str | None:
    for value in row[index + 1 :]:
        text = _clean_text(value)
        if text:
            return text
    return None


def _header_by_column(row: tuple[Any, ...]) -> dict[str, int | None] | None:
    header_by_field: dict[str, int] = {}
    for index, value in enumerate(row):
        field_name = _COLUMN_ALIASES.get(_normalize_text(value))
        if field_name is not None:
            header_by_field[field_name] = index

    if not all(column in header_by_field for column in _REQUIRED_COLUMNS):
        return None

    return {
        field_name: header_by_field.get(field_name)
        for field_name in (*_REQUIRED_COLUMNS, *_OPTIONAL_COLUMNS)
    }


def _parse_entry_row(
    row: tuple[Any, ...],
    conta_origem: str | None,
    header_by_column: dict[str, int | None],
) -> dict[str, Any] | None:
    data = _cell(row, header_by_column["data"])
    numero = _cell(row, header_by_column["numero"])
    row_conta_origem = _cell(row, header_by_column["conta_origem"])
    historico = _cell(row, header_by_column["historico"])
    contrapartida = _cell(row, header_by_column["contrapartida"])
    debito = _cell(row, header_by_column["debito"])
    credito = _cell(row, header_by_column["credito"])
    resolved_conta_origem = conta_origem or _clean_text(row_conta_origem)

    if (
        _is_blank_value(data)
        or _is_blank_value(historico)
        or _is_blank_value(resolved_conta_origem)
    ):
        return None

    return {
        "conta_origem": _clean_integer_like_text(resolved_conta_origem),
        "data": _format_entry_date(data),
        "numero": _clean_integer_like_text(numero),
        "historico": _clean_text(historico),
        "contrapartida": _clean_integer_like_text(contrapartida),
        "debito": debito if not _is_blank_value(debito) else None,
        "credito": credito if not _is_blank_value(credito) else None,
    }


def _cell(row: tuple[Any, ...], index: int | None) -> Any:
    if index is None or index >= len(row):
        return None
    return row[index]


def _build_metadata(
    empresa_nome: str | None,
    cnpj_cpf: str | None,
    periodo_inicio: str | None,
    periodo_fim: str | None,
    *,
    require_metadata: bool,
) -> RazaoMetadata:
    if require_metadata:
        missing_fields = [
            field_name
            for field_name, value in {
                "empresa_nome": empresa_nome,
                "cnpj_cpf": cnpj_cpf,
                "periodo_inicio": periodo_inicio,
                "periodo_fim": periodo_fim,
            }.items()
            if _is_blank_value(value)
        ]
        if missing_fields:
            raise RazaoParseError(
                "Cabecalho do razao sem metadados obrigatorios: "
                + ", ".join(missing_fields)
                + "."
            )

    return RazaoMetadata(
        empresa_nome=empresa_nome or "",
        cnpj_cpf=cnpj_cpf or "",
        periodo_inicio=periodo_inicio or "",
        periodo_fim=periodo_fim or "",
    )


def _parse_period_range(value: Any) -> tuple[str | None, str | None]:
    text = _clean_text(value)
    if not text or " - " not in text:
        return None, None

    start_text, end_text = text.split(" - ", 1)
    return _parse_br_date(start_text), _parse_br_date(end_text)


def _parse_br_date(value: Any) -> str | None:
    text = _clean_text(value)
    if not text:
        return None
    if isinstance(value, datetime):
        parsed_date = value.date()
    elif isinstance(value, date):
        parsed_date = value
    else:
        try:
            day, month, year = text.split("/")
            parsed_date = date(int(year), int(month), int(day))
        except ValueError:
            return None
    return parsed_date.isoformat()


def _format_entry_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    parsed_br_date = _parse_br_date(value)
    if parsed_br_date is not None:
        return parsed_br_date
    return _clean_text(value)


def _only_digits(value: Any) -> str | None:
    text = _clean_text(value)
    if not text:
        return None
    return re.sub(r"\D", "", text)


def _is_balance_row(row: tuple[Any, ...]) -> bool:
    return any("saldo anterior" in _normalize_text(value) for value in row)


def _is_empty_row(row: tuple[Any, ...]) -> bool:
    return all(_is_blank_value(value) for value in row)


def _is_blank_value(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _clean_text(value: Any) -> str | None:
    if _is_blank_value(value):
        return None
    return str(value).strip()


def _clean_integer_like_text(value: Any) -> str | None:
    text = _clean_text(value)
    if text is None:
        return None
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    return text


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return (
        str(value)
        .strip()
        .lower()
        .replace("ç", "c")
        .replace("ã", "a")
        .replace("á", "a")
        .replace("â", "a")
        .replace("é", "e")
        .replace("ê", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ô", "o")
        .replace("ú", "u")
    )

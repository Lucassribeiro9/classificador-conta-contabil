from datetime import date, datetime, timedelta, timezone
from io import BytesIO

import jwt
import pytest
from openpyxl import Workbook, load_workbook
from pwdlib import PasswordHash

from core.config import settings
from core.service_credentials import emitir_credencial_servico
from core.models import (
    AuditEvent,
    ContaContabil,
    Empresa,
    EmpresaContaContabil,
    IdentidadeServico,
    IdentidadeServicoEmpresa,
    IdentidadeServicoEscopo,
    LoteImportacaoMovimentoOperacional,
    MovimentoOperacionalImportado,
    Usuario,
    UsuarioEmpresaPermissao,
)


password_hash = PasswordHash.recommended()


@pytest.fixture(autouse=True)
def jwt_settings():
    previous_secret = settings.JWT_SECRET_KEY
    previous_algorithm = settings.JWT_ALGORITHM
    settings.JWT_SECRET_KEY = "test-secret"
    settings.JWT_ALGORITHM = "HS256"
    try:
        yield
    finally:
        settings.JWT_SECRET_KEY = previous_secret
        settings.JWT_ALGORITHM = previous_algorithm


@pytest.fixture(autouse=True)
def service_credential_settings():
    previous_secret = settings.SERVICE_CREDENTIAL_SECRET
    settings.SERVICE_CREDENTIAL_SECRET = "segredo-hmac-de-teste"
    try:
        yield
    finally:
        settings.SERVICE_CREDENTIAL_SECRET = previous_secret


def _usuario(**overrides) -> Usuario:
    data = {
        "nome": "Olivia Operadora",
        "login": "olivia.movimentos",
        "email": "olivia.movimentos@example.com",
        "senha_hash": password_hash.hash("senha-segura-123"),
        "papel": "operador",
        "is_active": True,
    }
    data.update(overrides)
    return Usuario(**data)


def _empresa(**overrides) -> Empresa:
    data = {
        "nome_empresa": "Empresa Movimentos LTDA",
        "cnpj_cpf": "11222333000144",
        "api_key": "api-key-movimentos",
        "cod_dominio": 1122,
    }
    data.update(overrides)
    return Empresa(**data)


def _conta(codigo: int) -> ContaContabil:
    return ContaContabil(
        codigo=codigo,
        classificacao=f"1.1.{codigo}",
        nome=f"CONTA {codigo}",
        tipo="A",
        grau=6,
        is_active=True,
    )


def _vinculo(empresa_id: int, conta_codigo: int) -> EmpresaContaContabil:
    return EmpresaContaContabil(
        empresa_id=empresa_id,
        conta_codigo=conta_codigo,
        quantidade_lancamentos=1,
        ultima_utilizacao=date(2026, 1, 1),
    )


def _access_token(usuario: Usuario) -> str:
    now = datetime.now(timezone.utc)
    return jwt.encode(
        {
            "sub": str(usuario.id),
            "role": usuario.papel,
            "type": "access",
            "iat": now,
            "exp": now + timedelta(hours=12),
        },
        settings.JWT_SECRET_KEY,
        algorithm=settings.JWT_ALGORITHM,
    )


def _auth_headers(usuario: Usuario) -> dict[str, str]:
    return {"Authorization": f"Bearer {_access_token(usuario)}"}


def _service_headers_for_empresa(
    empresa_id: int, *, escopo: str, status: str = "ativa"
) -> dict[str, str]:
    from tests.conftest import TestingSessionLocal

    with TestingSessionLocal() as session:
        empresa = session.get(Empresa, empresa_id)
        identidade = IdentidadeServico(
            identifier=f"n8n-{empresa.cod_dominio}-{escopo.split(':')[-1]}",
            nome="n8n Integracao",
            credential_hash="pendente",
            credential_fingerprint="pendente",
            status=status,
            empresas=[IdentidadeServicoEmpresa(empresa=empresa)],
            escopos=[IdentidadeServicoEscopo(escopo=escopo)],
        )
        session.add(identidade)
        session.commit()
        credencial = emitir_credencial_servico(session, identidade_id=identidade.id)
        if status != "ativa":
            identidade.status = status
        session.commit()
    return {"X-Service-Credential": credencial.secret}


def _movimentos_xlsx(cnpj_cpf: str = "11.222.333/0001-44") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Movimentos"
    sheet.append(["Empresa", "Empresa Movimentos LTDA"])
    sheet.append(["Codigo dominio", "1122"])
    sheet.append(["CNPJ/CPF", cnpj_cpf])
    sheet.append(["Periodo inicio", "01/01/2026"])
    sheet.append(["Periodo fim", "31/01/2026"])
    sheet.append([])
    sheet.append(
        [
            "data",
            "conta_financeira",
            "historico",
            "valor",
            "contrapartida",
            "tipo_movimento",
            "documento",
            "observacao",
        ]
    )
    sheet.append(
        [
            "02/01/2026",
            10046,
            "Pagamento fornecedor sensivel",
            -250.75,
            20001,
            "saida",
            "DOC-SENSIVEL-001",
            "Observacao sensivel",
        ]
    )

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    buffer.seek(0)
    return buffer.read()


def _upload_file(filename: str = "movimentos.xlsx") -> dict:
    return {
        "file": (
            filename,
            _movimentos_xlsx(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }


def _seed_user_company_catalog_and_links(permissao: str = "operacao"):
    from tests.conftest import TestingSessionLocal

    usuario = _usuario()
    empresa = _empresa()
    usuario.permissoes_empresas.append(
        UsuarioEmpresaPermissao(empresa=empresa, permissao=permissao)
    )
    contas = [_conta(10046), _conta(20001)]
    with TestingSessionLocal() as session:
        session.add_all([usuario, *contas])
        session.flush()
        session.add_all([_vinculo(empresa.id, 10046), _vinculo(empresa.id, 20001)])
        session.commit()
        session.refresh(usuario)
        session.refresh(empresa)
        return usuario, empresa.id


def _seed_operational_lote_with_movements(
    *,
    permissao: str = "leitura",
    empresa_overrides: dict | None = None,
    usuario_overrides: dict | None = None,
    lote_overrides: dict | None = None,
    movimentos: list[dict] | None = None,
):
    from tests.conftest import TestingSessionLocal

    empresa = _empresa(**(empresa_overrides or {}))
    usuario = _usuario(
        **(
            usuario_overrides
            or {
                "login": f"consulta.movimentos.{empresa.cod_dominio}",
                "email": f"consulta.movimentos.{empresa.cod_dominio}@example.com",
            }
        )
    )
    usuario.permissoes_empresas.append(
        UsuarioEmpresaPermissao(empresa=empresa, permissao=permissao)
    )
    lote_data = {
        "empresa": empresa,
        "usuario": usuario,
        "original_filename": "movimentos-consulta.xlsx",
        "file_hash": f"sha256:movimentos-consulta-{empresa.cod_dominio}",
        "status": "completed_with_warnings",
        "total_linhas": 2,
        "total_importadas": 2,
        "total_invalidas": 0,
        "warnings_metadata": {"warnings": []},
        "periodo_inicio": date(2026, 1, 1),
        "periodo_fim": date(2026, 1, 31),
        "cnpj_cpf_arquivo": empresa.cnpj_cpf,
        "codigo_dominio_arquivo": str(empresa.cod_dominio),
    }
    lote_data.update(lote_overrides or {})
    lote = LoteImportacaoMovimentoOperacional(**lote_data)
    movimentos_data = movimentos or [
        {
            "data": date(2026, 1, 2),
            "conta_financeira": 10046,
            "historico": "Pagamento fornecedor sensivel",
            "historico_normalizado": "pagamento fornecedor",
            "valor_original": -250.75,
            "valor_absoluto": 250.75,
            "direcao": "credito",
            "tipo_movimento": "saida",
            "documento": "DOC-SENSIVEL-001",
            "observacao": "Observacao sensivel",
            "contrapartida_informada": 20001,
            "status": "pre_classificado",
            "mensagens_validacao": [],
        },
        {
            "data": date(2026, 1, 3),
            "conta_financeira": 10046,
            "historico": "Transferencia sem contrapartida sensivel",
            "historico_normalizado": "transferencia sem contrapartida",
            "valor_original": -100,
            "valor_absoluto": 100,
            "direcao": "credito",
            "tipo_movimento": "transferencia",
            "documento": "DOC-SENSIVEL-002",
            "observacao": "Outra observacao sensivel",
            "contrapartida_informada": None,
            "status": "revisao",
            "mensagens_validacao": [
                "Tipo de movimento transferencia exige contrapartida."
            ],
        },
    ]

    with TestingSessionLocal() as session:
        session.add(lote)
        session.flush()
        for item in movimentos_data:
            session.add(
                MovimentoOperacionalImportado(
                    lote_id=lote.id,
                    empresa_id=empresa.id,
                    contrapartida_sugerida=None,
                    contrapartida_final=None,
                    confidence_sugerida=None,
                    elegivel_treino=False,
                    conta_debito=None,
                    conta_credito=None,
                    **item,
                )
            )
        session.commit()
        session.refresh(usuario)
        session.refresh(empresa)
        session.refresh(lote)
        return usuario, empresa.id, lote.id


def test_user_with_operacao_permission_imports_operational_movements(client):
    from tests.conftest import TestingSessionLocal

    usuario, empresa_id = _seed_user_company_catalog_and_links("operacao")

    response = client.post(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/import",
        files=_upload_file(),
        headers=_auth_headers(usuario),
    )

    assert response.status_code == 200
    assert response.json() == {
        "lote_id": 1,
        "status": "completed_with_warnings",
        "total_linhas": 1,
        "total_importadas": 1,
        "total_invalidas": 0,
        "warnings": [
            {
                "linha": 1,
                "warnings": [
                    "Saldo ausente; conferencia por saldo limitada para esta linha.",
                    "Saldo inicial ausente; saldo calculado partiu de zero.",
                ],
            }
        ],
    }

    with TestingSessionLocal() as session:
        lote = session.query(LoteImportacaoMovimentoOperacional).one()
        movimento = session.query(MovimentoOperacionalImportado).one()
        event = session.query(AuditEvent).one()
        assert lote.empresa_id == empresa_id
        assert movimento.empresa_id == empresa_id
        assert movimento.status == "pre_classificado"
        assert event.event_type == "operational_movements.imported"
        assert event.user_id == usuario.id
        assert event.empresa_id == empresa_id
        assert event.resource_id == str(lote.id)
        assert event.metadata_json["total_importadas"] == 1
        assert event.metadata_json["file_hash"].startswith("sha256:")
        assert "Pagamento fornecedor sensivel" not in str(event.metadata_json)
        assert "DOC-SENSIVEL-001" not in str(event.metadata_json)


def test_user_without_company_access_cannot_import_operational_movements(client):
    from tests.conftest import TestingSessionLocal

    usuario = _usuario(login="sem.acesso.movimentos", email="sem.acesso@example.com")
    empresa = _empresa()
    with TestingSessionLocal() as session:
        session.add_all([usuario, empresa, _conta(10046), _conta(20001)])
        session.commit()
        session.refresh(usuario)
        session.refresh(empresa)
        empresa_id = empresa.id

    response = client.post(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/import",
        files=_upload_file(),
        headers=_auth_headers(usuario),
    )

    assert response.status_code == 403
    assert response.json()["message"] == "Acesso negado"

    with TestingSessionLocal() as session:
        event = session.query(AuditEvent).one()
        assert event.event_type == "operational_movements.import_denied"
        assert event.user_id == usuario.id
        assert event.empresa_id == empresa_id
        assert event.metadata_json["reason"] == "access_denied"
        assert "Pagamento fornecedor sensivel" not in str(event.metadata_json)
        assert session.query(LoteImportacaoMovimentoOperacional).count() == 0
        assert session.query(MovimentoOperacionalImportado).count() == 0


def test_operational_movements_import_rejects_non_xlsx_file(client):
    from tests.conftest import TestingSessionLocal

    usuario, empresa_id = _seed_user_company_catalog_and_links("operacao")

    response = client.post(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/import",
        files={"file": ("movimentos.csv", b"data,historico", "text/csv")},
        headers=_auth_headers(usuario),
    )

    assert response.status_code == 400
    assert response.json()["message"] == "Arquivo deve ser .xlsx"

    with TestingSessionLocal() as session:
        event = session.query(AuditEvent).one()
        assert event.event_type == "operational_movements.import_failed"
        assert event.user_id == usuario.id
        assert event.empresa_id == empresa_id
        assert event.metadata_json["reason"] == "invalid_file_type"
        assert event.metadata_json["file_hash"].startswith("sha256:")
        assert "data,historico" not in str(event.metadata_json)


def test_user_with_leitura_permission_lists_own_operational_lotes(client):
    usuario, empresa_id, lote_id = _seed_operational_lote_with_movements(
        permissao="leitura"
    )
    _seed_operational_lote_with_movements(
        empresa_overrides={
            "nome_empresa": "Outra Empresa Movimentos LTDA",
            "cnpj_cpf": "99888777000166",
            "api_key": "api-key-outra-movimentos",
            "cod_dominio": 2211,
        },
        usuario_overrides={
            "login": "consulta.movimentos.outra",
            "email": "consulta.movimentos.outra@example.com",
        },
    )

    response = client.get(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/lotes",
        headers=_auth_headers(usuario),
    )

    assert response.status_code == 200
    assert response.json()["total"] == 1
    assert response.json()["page"] == 1
    assert response.json()["limit"] == 100
    assert response.json()["has_next"] is False
    assert response.json()["items"] == [
        {
            "id": lote_id,
            "empresa_id": empresa_id,
            "original_filename": "movimentos-consulta.xlsx",
            "status": "completed_with_warnings",
            "total_linhas": 2,
            "total_importadas": 2,
            "total_invalidas": 0,
            "periodo_inicio": "2026-01-01",
            "periodo_fim": "2026-01-31",
            "layout_version": "operacional_valor_legado_v1",
            "created_at": response.json()["items"][0]["created_at"],
        }
    ]


def test_user_with_leitura_permission_downloads_classified_operational_sheet(client):
    from tests.conftest import TestingSessionLocal

    usuario, empresa_id, lote_id = _seed_operational_lote_with_movements(
        permissao="leitura"
    )

    response = client.get(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/lotes/{lote_id}/planilha-classificada",
        headers=_auth_headers(usuario),
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["content-disposition"] == (
        f'attachment; filename="11222333000144-lote{lote_id}-classificada.xlsx"'
    )

    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook["Movimentos"]
    headers = [cell.value for cell in sheet[1]]
    assert "lote_id" in headers
    assert "movimento_id" in headers
    assert "export_revision" in headers
    assert "row_version" in headers
    assert "decisao_revisao" in headers
    assert "contrapartida_final" in headers
    assert "observacao_revisao" in headers
    assert sheet.max_row == 3

    with TestingSessionLocal() as session:
        event = session.query(AuditEvent).one()

    assert event.event_type == "operational_movements.classified_sheet_downloaded"
    assert event.user_id == usuario.id
    assert event.empresa_id == empresa_id
    assert event.resource_id == str(lote_id)
    assert event.metadata_json["lote_id"] == lote_id
    assert event.metadata_json["layout_version"] == "operacional_valor_legado_v1"
    assert event.metadata_json["total_movimentos"] == 2
    assert event.metadata_json["export_revision"]
    assert "Pagamento fornecedor sensivel" not in str(event.metadata_json)
    assert "DOC-SENSIVEL-001" not in str(event.metadata_json)


def test_service_with_download_scope_downloads_classified_operational_sheet(client):
    from tests.conftest import TestingSessionLocal

    _usuario, empresa_id, lote_id = _seed_operational_lote_with_movements(
        permissao="leitura"
    )
    headers = _service_headers_for_empresa(empresa_id, escopo="movimentos:download")

    response = client.get(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/lotes/{lote_id}/planilha-classificada",
        headers=headers,
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = load_workbook(BytesIO(response.content))
    assert workbook["Movimentos"].max_row == 3

    with TestingSessionLocal() as session:
        event = session.query(AuditEvent).filter_by(
            event_type="operational_movements.classified_sheet_downloaded"
        ).one()

    assert event.user_id is None
    assert event.empresa_id == empresa_id
    assert event.resource_id == str(lote_id)
    assert event.metadata_json["actor_type"] == "service"
    assert event.metadata_json["scope"] == "movimentos:download"
    assert event.metadata_json["identidade_servico_id"]
    assert event.metadata_json["identifier"].startswith("n8n-")
    assert event.metadata_json["credential_fingerprint"].startswith("fp_")
    assert "X-Service-Credential" not in str(event.metadata_json)
    assert headers["X-Service-Credential"] not in str(event.metadata_json)


def test_classified_operational_sheet_requires_jwt_not_api_or_admin_token(client):
    _, empresa_id, lote_id = _seed_operational_lote_with_movements(permissao="leitura")

    for headers in (
        {"X-API-Key": "api-key-movimentos"},
        {"X-Admin-Token": "test-admin-token"},
    ):
        response = client.get(
            f"/api/v1/companies/{empresa_id}/movimentos-operacionais/lotes/{lote_id}/planilha-classificada",
            headers=headers,
        )

        assert response.status_code == 401


def test_user_without_company_access_cannot_download_classified_operational_sheet(
    client,
):
    from tests.conftest import TestingSessionLocal

    usuario = _usuario(
        login="sem.download.movimentos",
        email="sem.download.movimentos@example.com",
    )
    _, empresa_id, lote_id = _seed_operational_lote_with_movements(permissao="leitura")
    with TestingSessionLocal() as session:
        session.add(usuario)
        session.commit()
        session.refresh(usuario)

    response = client.get(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/lotes/{lote_id}/planilha-classificada",
        headers=_auth_headers(usuario),
    )

    assert response.status_code == 403
    assert response.json()["message"] == "Acesso negado"


def test_classified_operational_sheet_does_not_cross_company_boundaries(client):
    usuario, empresa_id, _ = _seed_operational_lote_with_movements(permissao="leitura")
    _, outra_empresa_id, outro_lote_id = _seed_operational_lote_with_movements(
        empresa_overrides={
            "nome_empresa": "Empresa Download Sem Acesso LTDA",
            "cnpj_cpf": "77666555000144",
            "api_key": "api-key-download-sem-acesso",
            "cod_dominio": 4411,
        },
        usuario_overrides={
            "login": "download.movimentos.sem.acesso",
            "email": "download.movimentos.sem.acesso@example.com",
        },
    )

    response = client.get(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/lotes/{outro_lote_id}/planilha-classificada",
        headers=_auth_headers(usuario),
    )

    assert outra_empresa_id != empresa_id
    assert response.status_code == 404
    assert response.json()["message"] == "Lote operacional não encontrado"


def test_global_admin_can_download_classified_operational_sheet_without_company_link(
    client,
):
    from tests.conftest import TestingSessionLocal

    _, empresa_id, lote_id = _seed_operational_lote_with_movements(permissao="leitura")
    admin = _usuario(
        login="admin.download.movimentos",
        email="admin.download.movimentos@example.com",
        papel="admin",
    )
    with TestingSessionLocal() as session:
        session.add(admin)
        session.commit()
        session.refresh(admin)

    response = client.get(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/lotes/{lote_id}/planilha-classificada",
        headers=_auth_headers(admin),
    )

    assert response.status_code == 200
    assert response.headers["content-disposition"] == (
        f'attachment; filename="11222333000144-lote{lote_id}-classificada.xlsx"'
    )


def test_classified_operational_sheet_rejects_unknown_layout(client):
    usuario, empresa_id, lote_id = _seed_operational_lote_with_movements(
        permissao="leitura",
        lote_overrides={"layout_version": "operacional_desconhecido_v99"},
    )

    response = client.get(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/lotes/{lote_id}/planilha-classificada",
        headers=_auth_headers(usuario),
    )

    assert response.status_code == 400
    assert response.json()["message"] == (
        "Layout operacional desconhecido: operacional_desconhecido_v99"
    )


def test_user_lists_operational_movements_by_lote_and_status_without_raw_payload(
    client,
):
    usuario, empresa_id, lote_id = _seed_operational_lote_with_movements(
        permissao="leitura",
        movimentos=[
            {
                "data": date(2026, 1, 3),
                "conta_financeira": 10046,
                "historico": "Transferencia sem contrapartida sensivel",
                "historico_normalizado": "transferencia sem contrapartida",
                "valor_original": -100,
                "valor_absoluto": 100,
                "direcao": "credito",
                "tipo_movimento": "transferencia",
                "documento": "DOC-SENSIVEL-002",
                "observacao": "Outra observacao sensivel",
                "contrapartida_informada": None,
                "status": "revisao",
                "mensagens_validacao": [
                    "Tipo de movimento transferencia exige contrapartida."
                ],
                "saldo_observado_original": "saldo digitado manualmente",
                "saldo_observado_decimal": None,
                "saldo_calculado_decimal": 1500,
                "warnings_saldo": [
                    "Saldo informado invalido; conferencia por saldo limitada para esta linha."
                ],
            }
        ],
    )

    response = client.get(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/lotes/{lote_id}/movimentos",
        params={"status": "revisao"},
        headers=_auth_headers(usuario),
    )

    assert response.status_code == 200
    assert response.json()["total"] == 1
    assert response.json()["items"] == [
        {
            "id": response.json()["items"][0]["id"],
            "lote_id": lote_id,
            "empresa_id": empresa_id,
            "data": "2026-01-03",
            "conta_financeira": 10046,
            "historico_normalizado": "transferencia sem contrapartida",
            "valor_absoluto": "100.00",
            "direcao": "credito",
            "tipo_movimento": "transferencia",
            "saldo_observado_original": None,
            "saldo_observado_decimal": None,
            "saldo_calculado_decimal": "1500.00",
            "warnings_saldo": [
                "Saldo informado invalido; conferencia por saldo limitada para esta linha."
            ],
            "contrapartida_informada": None,
            "contrapartida_sugerida": None,
            "contrapartida_final": None,
            "confidence_sugerida": None,
            "status": "revisao",
            "elegivel_treino": False,
            "mensagens_validacao": [
                "Tipo de movimento transferencia exige contrapartida."
            ],
            "conta_debito": None,
            "conta_credito": None,
        }
    ]
    item = response.json()["items"][0]
    assert "historico" not in item
    assert "documento" not in item
    assert "observacao" not in item
    assert "DOC-SENSIVEL-002" not in str(response.json())
    assert "saldo digitado manualmente" not in str(response.json())


def test_user_without_company_access_cannot_list_operational_lotes(client):
    from tests.conftest import TestingSessionLocal

    usuario = _usuario(login="sem.consulta.mov", email="sem.consulta.mov@example.com")
    _, empresa_id, _ = _seed_operational_lote_with_movements(permissao="leitura")
    with TestingSessionLocal() as session:
        session.add(usuario)
        session.commit()
        session.refresh(usuario)

    response = client.get(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/lotes",
        headers=_auth_headers(usuario),
    )

    assert response.status_code == 403
    assert response.json()["message"] == "Acesso negado"


def test_operational_movements_query_does_not_cross_company_boundaries(client):
    usuario, empresa_id, _ = _seed_operational_lote_with_movements(permissao="leitura")
    _, outra_empresa_id, outro_lote_id = _seed_operational_lote_with_movements(
        empresa_overrides={
            "nome_empresa": "Empresa Sem Acesso Movimentos LTDA",
            "cnpj_cpf": "88777666000155",
            "api_key": "api-key-sem-acesso-movimentos",
            "cod_dominio": 3311,
        },
        usuario_overrides={
            "login": "consulta.movimentos.sem.acesso",
            "email": "consulta.movimentos.sem.acesso@example.com",
        },
    )

    response = client.get(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/lotes/{outro_lote_id}/movimentos",
        headers=_auth_headers(usuario),
    )

    assert outra_empresa_id != empresa_id
    assert response.status_code == 404
    assert response.json()["message"] == "Lote operacional não encontrado"


def test_user_with_operacao_permission_classifies_pending_operational_movements(
    client,
    monkeypatch,
):
    usuario, empresa_id, _ = _seed_operational_lote_with_movements(permissao="operacao")

    def fake_classificar(db, *, empresa_id, model_dir=None):
        return {
            "empresa_id": empresa_id,
            "quantidade_processada": 2,
            "total_sugerido": 1,
            "total_revisao": 1,
        }

    monkeypatch.setattr(
        "api.routes.movimentos_operacionais.classificar_movimentos_operacionais_pendentes",
        fake_classificar,
    )

    response = client.post(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/classificar",
        headers=_auth_headers(usuario),
    )

    assert response.status_code == 200
    assert response.json() == {
        "empresa_id": empresa_id,
        "quantidade_processada": 2,
        "total_sugerido": 1,
        "total_revisao": 1,
    }


def test_classificar_movimentos_returns_controlled_error_when_model_is_missing(
    client,
    monkeypatch,
    tmp_path,
):
    from tests.conftest import TestingSessionLocal

    usuario, empresa_id, _ = _seed_operational_lote_with_movements(
        permissao="operacao",
        movimentos=[
            {
                "data": date(2026, 1, 3),
                "conta_financeira": 10046,
                "historico": "Pagamento pendente sensivel",
                "historico_normalizado": "pagamento pendente",
                "valor_original": -100,
                "valor_absoluto": 100,
                "direcao": "credito",
                "tipo_movimento": "saida",
                "documento": "DOC-SENSIVEL-MODELO-AUSENTE",
                "observacao": "Nao deve ir para auditoria",
                "contrapartida_informada": None,
                "status": "pendente",
                "mensagens_validacao": [],
            },
        ],
    )
    monkeypatch.setattr(settings, "MODEL_DIR", str(tmp_path))

    response = client.post(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/classificar",
        headers=_auth_headers(usuario),
    )

    assert response.status_code == 404
    assert response.json()["message"] == "Modelo treinado não encontrado para a empresa"

    with TestingSessionLocal() as session:
        event = session.query(AuditEvent).one()

    assert event.event_type == "operational_movements.classification_failed"
    assert event.user_id == usuario.id
    assert event.empresa_id == empresa_id
    assert event.resource_id == "operational_movements_classification"
    assert event.metadata_json == {
        "total_pendentes": 1,
        "error_type": "ModelNotFound",
        "reason": "model_not_found",
    }
    assert "Pagamento pendente sensivel" not in str(event.metadata_json)
    assert "DOC-SENSIVEL-MODELO-AUSENTE" not in str(event.metadata_json)


def _feedback_roundtrip_xlsx(rows: list[dict]) -> bytes:
    headers = [
        "lote_id",
        "movimento_id",
        "linha_original",
        "layout_version",
        "export_revision",
        "row_version",
        "contrapartida_sugerida",
        "confidence_sugerida",
        "status_atual",
        "mensagem_validacao",
        "decisao_revisao",
        "contrapartida_final",
        "observacao_revisao",
    ]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Movimentos"
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header) for header in headers])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return output.read()


def test_import_classified_sheet_feedback_endpoint_applies_partial_file(client):
    from tests.conftest import TestingSessionLocal

    usuario, empresa_id, lote_id = _seed_operational_lote_with_movements(
        permissao="operacao"
    )

    with TestingSessionLocal() as session:
        session.add(
            ContaContabil(
                codigo=20001,
                classificacao="2.0.0",
                nome="Conta 20001",
                tipo="A",
                grau=3,
            )
        )
        session.commit()
        movimentos = (
            session.query(MovimentoOperacionalImportado)
            .filter_by(lote_id=lote_id)
            .order_by(MovimentoOperacionalImportado.id.asc())
            .all()
        )
        movimentos[0].status = "sugerido"
        movimentos[0].contrapartida_sugerida = 20001
        movimentos[0].confidence_sugerida = 0.91
        session.commit()
        rows = [
            {
                "lote_id": lote_id,
                "movimento_id": movimentos[0].id,
                "linha_original": movimentos[0].linha_original,
                "layout_version": "operacional_valor_legado_v1",
                "export_revision": "revision-api",
                "row_version": movimentos[0].row_version,
                "contrapartida_sugerida": 20001,
                "confidence_sugerida": 0.91,
                "status_atual": "sugerido",
                "decisao_revisao": "aprovar",
                "contrapartida_final": 20001,
            },
            {
                "lote_id": lote_id,
                "movimento_id": movimentos[1].id,
                "linha_original": movimentos[1].linha_original,
                "layout_version": "operacional_valor_legado_v1",
                "export_revision": "revision-api",
                "row_version": movimentos[1].row_version,
                "status_atual": movimentos[1].status,
                "decisao_revisao": "",
            },
        ]

    response = client.post(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/lotes/{lote_id}/planilha-classificada/feedback",
        files={
            "file": (
                "feedback.xlsx",
                _feedback_roundtrip_xlsx(rows),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=_auth_headers(usuario),
    )

    assert response.status_code == 200
    assert response.json()["total_linhas"] == 2
    assert response.json()["total_aplicado"] == 1
    assert response.json()["total_ignorado"] == 1
    assert [item["status"] for item in response.json()["resultados"]] == [
        "aplicada",
        "ignorada",
    ]

    with TestingSessionLocal() as session:
        movimento = session.get(MovimentoOperacionalImportado, movimentos[0].id)
        event_types = [event.event_type for event in session.query(AuditEvent).all()]

    assert movimento.status == "aprovado"
    assert movimento.contrapartida_final == 20001
    assert "operational_movements.feedback_imported" in event_types


def test_roundtrip_rejects_ambiguous_human_and_service_credentials(client):
    usuario, empresa_id, lote_id = _seed_operational_lote_with_movements(
        permissao="leitura"
    )
    service_headers = _service_headers_for_empresa(
        empresa_id, escopo="movimentos:download"
    )

    response = client.get(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/lotes/{lote_id}/planilha-classificada",
        headers={**_auth_headers(usuario), **service_headers},
    )

    assert response.status_code == 400
    assert response.json()["message"] == "Credenciais ambíguas"


def test_service_scope_is_independent_between_download_and_feedback(client):
    _usuario, empresa_id, lote_id = _seed_operational_lote_with_movements(
        permissao="operacao"
    )
    download_headers = _service_headers_for_empresa(
        empresa_id, escopo="movimentos:download"
    )
    feedback_headers = _service_headers_for_empresa(
        empresa_id, escopo="movimentos:feedback"
    )

    feedback_response = client.post(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/lotes/{lote_id}/planilha-classificada/feedback",
        files={
            "file": (
                "feedback.xlsx",
                _feedback_roundtrip_xlsx([]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=download_headers,
    )
    download_response = client.get(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/lotes/{lote_id}/planilha-classificada",
        headers=feedback_headers,
    )

    assert feedback_response.status_code == 403
    assert download_response.status_code == 403


def test_service_cross_company_is_blocked_and_audited(client):
    from tests.conftest import TestingSessionLocal

    _usuario, empresa_id, lote_id = _seed_operational_lote_with_movements(
        permissao="leitura"
    )
    _outro_usuario, outra_empresa_id, _outro_lote_id = _seed_operational_lote_with_movements(
        empresa_overrides={
            "nome_empresa": "Outra Empresa Integracao LTDA",
            "cnpj_cpf": "66555444000133",
            "api_key": "api-key-outra-integracao",
            "cod_dominio": 6655,
        },
        usuario_overrides={
            "login": "outra.integracao",
            "email": "outra.integracao@example.com",
        },
    )
    headers = _service_headers_for_empresa(
        outra_empresa_id, escopo="movimentos:download"
    )

    response = client.get(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/lotes/{lote_id}/planilha-classificada",
        headers=headers,
    )

    assert response.status_code == 403
    with TestingSessionLocal() as session:
        event = session.query(AuditEvent).filter_by(
            event_type="operational_movements.service_access_denied"
        ).one()

    assert event.user_id is None
    assert event.empresa_id == empresa_id
    assert event.metadata_json["actor_type"] == "service"
    assert event.metadata_json["reason"] == "access_denied"
    assert event.metadata_json["scope"] == "movimentos:download"
    assert headers["X-Service-Credential"] not in str(event.metadata_json)


def test_service_lote_inexistente_is_blocked_and_audited(client):
    from tests.conftest import TestingSessionLocal

    _usuario, empresa_id, _lote_id = _seed_operational_lote_with_movements(
        permissao="leitura"
    )
    headers = _service_headers_for_empresa(empresa_id, escopo="movimentos:download")

    response = client.get(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/lotes/9999/planilha-classificada",
        headers=headers,
    )

    assert response.status_code == 404
    with TestingSessionLocal() as session:
        event = session.query(AuditEvent).filter_by(
            event_type="operational_movements.service_access_denied"
        ).one()

    assert event.user_id is None
    assert event.empresa_id == empresa_id
    assert event.resource_id == "9999"
    assert event.metadata_json["reason"] == "lote_not_found"
    assert event.metadata_json["scope"] == "movimentos:download"
    assert headers["X-Service-Credential"] not in str(event.metadata_json)


def test_revoked_service_identity_is_blocked_and_audited(client):
    from tests.conftest import TestingSessionLocal

    _usuario, empresa_id, lote_id = _seed_operational_lote_with_movements(
        permissao="leitura"
    )
    headers = _service_headers_for_empresa(
        empresa_id, escopo="movimentos:download", status="revogada"
    )

    response = client.get(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/lotes/{lote_id}/planilha-classificada",
        headers=headers,
    )

    assert response.status_code == 403
    with TestingSessionLocal() as session:
        event = session.query(AuditEvent).filter_by(
            event_type="operational_movements.service_access_denied"
        ).one()

    assert event.metadata_json["reason"] == "inactive_or_revoked"
    assert event.metadata_json["credential_fingerprint"].startswith("fp_")
    assert headers["X-Service-Credential"] not in str(event.metadata_json)


def test_service_with_feedback_scope_imports_classified_sheet_feedback(client):
    from tests.conftest import TestingSessionLocal

    _usuario, empresa_id, lote_id = _seed_operational_lote_with_movements(
        permissao="operacao"
    )
    headers = _service_headers_for_empresa(empresa_id, escopo="movimentos:feedback")

    with TestingSessionLocal() as session:
        session.add(_conta(20001))
        session.commit()
        movimentos = (
            session.query(MovimentoOperacionalImportado)
            .filter_by(lote_id=lote_id)
            .order_by(MovimentoOperacionalImportado.id.asc())
            .all()
        )
        movimentos[0].status = "sugerido"
        movimentos[0].contrapartida_sugerida = 20001
        movimentos[0].confidence_sugerida = 0.91
        session.commit()
        movimento_id = movimentos[0].id
        rows = [
            {
                "lote_id": lote_id,
                "movimento_id": movimentos[0].id,
                "linha_original": movimentos[0].linha_original,
                "layout_version": "operacional_valor_legado_v1",
                "export_revision": "revision-service-api",
                "row_version": movimentos[0].row_version,
                "contrapartida_sugerida": 20001,
                "confidence_sugerida": 0.91,
                "status_atual": "sugerido",
                "decisao_revisao": "aprovar",
                "contrapartida_final": 20001,
            }
        ]

    response = client.post(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/lotes/{lote_id}/planilha-classificada/feedback",
        files={
            "file": (
                "feedback.xlsx",
                _feedback_roundtrip_xlsx(rows),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=headers,
    )

    assert response.status_code == 200
    assert response.json()["total_aplicado"] == 1

    with TestingSessionLocal() as session:
        movimento = session.get(MovimentoOperacionalImportado, movimento_id)
        imported_event = (
            session.query(AuditEvent)
            .filter_by(event_type="operational_movements.feedback_imported")
            .one()
        )
        decision_event = (
            session.query(AuditEvent)
            .filter_by(
                event_type="operational_movements.aprovado",
                resource_id=str(movimento_id),
            )
            .one()
        )

    assert movimento.status == "aprovado"
    assert imported_event.user_id is None
    assert imported_event.metadata_json["actor_type"] == "service"
    assert imported_event.metadata_json["scope"] == "movimentos:feedback"
    assert decision_event.user_id is None
    assert decision_event.metadata_json["actor_type"] == "service"
    assert headers["X-Service-Credential"] not in str(imported_event.metadata_json)


def test_service_feedback_lote_inexistente_is_blocked_and_audited(client):
    from tests.conftest import TestingSessionLocal

    _usuario, empresa_id, _lote_id = _seed_operational_lote_with_movements(
        permissao="operacao"
    )
    headers = _service_headers_for_empresa(empresa_id, escopo="movimentos:feedback")

    response = client.post(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/lotes/9999/planilha-classificada/feedback",
        files={
            "file": (
                "feedback.xlsx",
                _feedback_roundtrip_xlsx([]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=headers,
    )

    assert response.status_code == 404
    with TestingSessionLocal() as session:
        event = session.query(AuditEvent).filter_by(
            event_type="operational_movements.service_access_denied"
        ).one()

    assert event.user_id is None
    assert event.empresa_id == empresa_id
    assert event.resource_id == "9999"
    assert event.metadata_json["reason"] == "lote_not_found"
    assert event.metadata_json["scope"] == "movimentos:feedback"
    assert headers["X-Service-Credential"] not in str(event.metadata_json)


def test_import_classified_sheet_feedback_returns_400_for_structural_error(client):
    usuario, empresa_id, lote_id = _seed_operational_lote_with_movements(
        permissao="operacao"
    )
    workbook = Workbook()
    workbook.active.title = "OutraAba"
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)

    response = client.post(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/lotes/{lote_id}/planilha-classificada/feedback",
        files={
            "file": (
                "feedback.xlsx",
                output.read(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=_auth_headers(usuario),
    )

    assert response.status_code == 400
    assert response.json()["message"] == "Aba Movimentos não encontrada"


def test_import_classified_sheet_feedback_requires_operational_permission(client):
    usuario, empresa_id, lote_id = _seed_operational_lote_with_movements(
        permissao="leitura"
    )

    response = client.post(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/lotes/{lote_id}/planilha-classificada/feedback",
        files={
            "file": (
                "feedback.xlsx",
                _feedback_roundtrip_xlsx([]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=_auth_headers(usuario),
    )

    assert response.status_code == 403
    assert response.json()["message"] == "Permissão insuficiente"


def test_review_movimento_approve_success(client):
    from tests.conftest import TestingSessionLocal
    from core.models import ContaContabil
    usuario, empresa_id, lote_id = _seed_operational_lote_with_movements(permissao="operacao")
    
    with TestingSessionLocal() as session:
        conta = ContaContabil(codigo=20001, classificacao="2.0.0", nome="Conta 20001", tipo="A", grau=3)
        session.add(conta)
        session.commit()
        mov = session.query(MovimentoOperacionalImportado).filter_by(lote_id=lote_id).first()
        mov_id = mov.id

    response = client.post(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/lotes/{lote_id}/movimentos/{mov_id}/review",
        json={"action": "approve", "conta_final": 20001},
        headers=_auth_headers(usuario),
    )

    assert response.status_code == 200
    assert response.json()["status"] == "aprovado"
    assert response.json()["contrapartida_final"] == 20001


def test_review_movimento_reject_success(client):
    from tests.conftest import TestingSessionLocal
    usuario, empresa_id, lote_id = _seed_operational_lote_with_movements(permissao="operacao")
    
    with TestingSessionLocal() as session:
        mov = session.query(MovimentoOperacionalImportado).filter_by(lote_id=lote_id).first()
        mov_id = mov.id

    response = client.post(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/lotes/{lote_id}/movimentos/{mov_id}/review",
        json={"action": "reject"},
        headers=_auth_headers(usuario),
    )

    assert response.status_code == 200
    assert response.json()["status"] == "rejeitado"
    assert response.json()["contrapartida_final"] is None


def test_review_movimento_rejects_finalized_status(client):
    from tests.conftest import TestingSessionLocal
    from core.models import ContaContabil

    usuario, empresa_id, lote_id = _seed_operational_lote_with_movements(
        permissao="operacao"
    )

    with TestingSessionLocal() as session:
        conta = ContaContabil(
            codigo=20001,
            classificacao="2.0.0",
            nome="Conta 20001",
            tipo="A",
            grau=3,
        )
        session.add(conta)
        session.commit()
        mov = (
            session.query(MovimentoOperacionalImportado)
            .filter_by(lote_id=lote_id)
            .first()
        )
        mov.status = "aprovado"
        mov.contrapartida_final = 20001
        session.commit()
        mov_id = mov.id

    response = client.post(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/lotes/{lote_id}/movimentos/{mov_id}/review",
        json={"action": "correct", "conta_final": 20001},
        headers=_auth_headers(usuario),
    )

    assert response.status_code == 400
    assert response.json()["message"] == "Status não permite revisão"


def test_review_movimento_without_permission(client):
    from tests.conftest import TestingSessionLocal
    usuario, empresa_id, lote_id = _seed_operational_lote_with_movements(permissao="leitura")
    
    with TestingSessionLocal() as session:
        mov = session.query(MovimentoOperacionalImportado).filter_by(lote_id=lote_id).first()
        mov_id = mov.id

    response = client.post(
        f"/api/v1/companies/{empresa_id}/movimentos-operacionais/lotes/{lote_id}/movimentos/{mov_id}/review",
        json={"action": "approve", "conta_final": 20001},
        headers=_auth_headers(usuario),
    )

    assert response.status_code == 403
    assert response.json()["message"] == "Permissão insuficiente"

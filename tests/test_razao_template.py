from pathlib import Path

from openpyxl import load_workbook


TEMPLATE_PATH = Path("modelo-razao-importacao.xlsx")


def test_modelo_razao_importacao_define_campos_esperados():
    workbook = load_workbook(TEMPLATE_PATH, read_only=True, data_only=True)
    try:
        sheet = workbook["Razao"]

        assert sheet["A3"].value == "Empresa"
        assert sheet["B3"].value == "EMPRESA MODELO LTDA"
        assert sheet["A4"].value == "CNPJ"
        assert sheet["B4"].value == "11.222.333/0001-81"
        assert sheet["A5"].value == "Periodo inicio"
        assert sheet["A6"].value == "Periodo fim"

        headers = [sheet.cell(10, column).value for column in range(1, 11)]
        assert headers == [
            "data",
            "numero",
            "conta_origem",
            "conta_origem_classificacao",
            "conta_origem_nome",
            "historico",
            "contrapartida",
            "debito",
            "credito",
            "saldo_exercicio_original",
        ]

        help_text = " ".join(
            str(sheet.cell(11, column).value) for column in range(1, 11)
        ).lower()
        assert "pode ficar vazio" in help_text
        assert "opcional" in help_text
        assert "debito ou credito" in sheet["B115"].value.lower()
        assert "cod_dominio" not in help_text
    finally:
        workbook.close()

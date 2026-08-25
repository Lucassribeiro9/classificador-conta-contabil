import asyncio
import os
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from io import BytesIO
from uuid import uuid4

import httpx
import jwt
import pytest
from openpyxl import load_workbook
from pwdlib import PasswordHash
from sqlalchemy import create_engine
from sqlalchemy.engine import make_url
from sqlalchemy.orm import sessionmaker

from api.dependencies import get_db
from api.main import app
from core.config import settings
from core.models import (
    AuditEvent,
    ContaContabil,
    Empresa,
    IdentidadeServico,
    IdentidadeServicoEmpresa,
    IdentidadeServicoEscopo,
    LoteImportacaoMovimentoOperacional,
    MovimentoOperacionalImportado,
    Usuario,
    UsuarioEmpresaPermissao,
)
from core.service_credentials import emitir_credencial_servico


pytestmark = pytest.mark.integration_postgres

XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
PASSWORD_HASH = PasswordHash.recommended()


class ASGITestClient:
    """Cliente síncrono para exercitar a API via ASGI sem servidor externo."""

    def __init__(self, target_app):
        self.app = target_app

    def request(self, method: str, url: str, **kwargs):
        async def send_request():
            transport = httpx.ASGITransport(app=self.app)
            async with httpx.AsyncClient(
                transport=transport,
                base_url="http://testserver",
            ) as client:
                return await client.request(method, url, **kwargs)

        return asyncio.run(send_request())

    def get(self, url: str, **kwargs):
        return self.request("GET", url, **kwargs)

    def post(self, url: str, **kwargs):
        return self.request("POST", url, **kwargs)


@pytest.fixture(scope="session")
def postgres_session_factory():
    database_url = os.environ.get("DATABASE_URL")
    if not database_url:
        pytest.fail("DATABASE_URL deve estar definido para testes PostgreSQL.")
    if make_url(database_url).get_backend_name() != "postgresql":
        pytest.fail("DATABASE_URL deve apontar para PostgreSQL nos testes.")

    engine = create_engine(database_url)
    try:
        yield sessionmaker(autocommit=False, autoflush=False, bind=engine)
    finally:
        engine.dispose()


@pytest.fixture()
def api_client(postgres_session_factory):
    async def override_get_db():
        db = postgres_session_factory()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    try:
        yield ASGITestClient(app)
    finally:
        app.dependency_overrides.clear()


@pytest.fixture(autouse=True)
def auth_settings():
    previous_jwt_secret = settings.JWT_SECRET_KEY
    previous_jwt_algorithm = settings.JWT_ALGORITHM
    previous_service_secret = settings.SERVICE_CREDENTIAL_SECRET
    settings.JWT_SECRET_KEY = "test-secret-roundtrip"
    settings.JWT_ALGORITHM = "HS256"
    settings.SERVICE_CREDENTIAL_SECRET = "segredo-hmac-roundtrip-integracao"
    try:
        yield
    finally:
        settings.JWT_SECRET_KEY = previous_jwt_secret
        settings.JWT_ALGORITHM = previous_jwt_algorithm
        settings.SERVICE_CREDENTIAL_SECRET = previous_service_secret


def test_servico_consumidor_baixa_edita_reenvia_e_reenvia_sem_duplicar(
    api_client,
    postgres_session_factory,
):
    data = _seed_roundtrip_lote(
        postgres_session_factory,
        layout_version="operacional_valor_saldo_v1",
        actor_kind="service",
    )
    headers = _service_headers(
        postgres_session_factory,
        empresa_id=data["empresa_id"],
        escopos=("movimentos:download", "movimentos:feedback"),
    )

    download = api_client.get(data["download_url"], headers=headers)
    assert download.status_code == 200
    assert download.headers["content-type"] == XLSX_MEDIA_TYPE

    feedback_content = _edit_feedback_sheet(
        download.content,
        {
            "aprovar": {
                "movimento_id": data["movimentos"]["aprovar"],
                "contrapartida_final": data["contas"]["aprovacao"],
            },
            "corrigir": {
                "movimento_id": data["movimentos"]["corrigir"],
                "decisao_revisao": "corrigir",
                "contrapartida_final": data["contas"]["correcao"],
            },
            "rejeitar": {"movimento_id": data["movimentos"]["rejeitar"]},
            "invalida": {
                "movimento_id": data["movimentos"]["invalida"],
                "decisao_revisao": "decisao-inexistente",
            },
        },
    )

    response = _post_feedback(api_client, data["feedback_url"], feedback_content, headers)
    assert response.status_code == 200
    assert response.json()["total_linhas"] == 5
    assert response.json()["total_aplicado"] == 3
    assert response.json()["total_ignorado"] == 1
    assert response.json()["total_invalido"] == 1
    assert _statuses(response) == [
        "aplicada",
        "aplicada",
        "aplicada",
        "ignorada",
        "invalida",
    ]

    decision_events_before = _decision_event_count(
        postgres_session_factory, data["empresa_id"]
    )
    replay = _post_feedback(api_client, data["feedback_url"], feedback_content, headers)
    assert replay.status_code == 200
    assert replay.json()["total_aplicado"] == 0
    assert replay.json()["total_ignorado"] == 4
    assert replay.json()["total_invalido"] == 1
    assert _decision_event_count(postgres_session_factory, data["empresa_id"]) == (
        decision_events_before
    )

    with postgres_session_factory() as session:
        aprovado = session.get(
            MovimentoOperacionalImportado, data["movimentos"]["aprovar"]
        )
        corrigido = session.get(
            MovimentoOperacionalImportado, data["movimentos"]["corrigir"]
        )
        rejeitado = session.get(
            MovimentoOperacionalImportado, data["movimentos"]["rejeitar"]
        )
        assert aprovado.status == "aprovado"
        assert aprovado.contrapartida_final == data["contas"]["aprovacao"]
        assert corrigido.status == "corrigido"
        assert corrigido.contrapartida_final == data["contas"]["correcao"]
        assert rejeitado.status == "rejeitado"
        assert rejeitado.contrapartida_final is None

    _assert_audit_metadata_safe(
        postgres_session_factory,
        empresa_id=data["empresa_id"],
        forbidden=[headers["X-Service-Credential"], "Historico Tratado", "DOC-TESTE"],
    )


def test_usuario_jwt_consumidor_nao_sobrescreve_revisao_mais_recente(
    api_client,
    postgres_session_factory,
):
    data = _seed_roundtrip_lote(
        postgres_session_factory,
        layout_version="operacional_debito_credito_saldo_v1",
        actor_kind="user",
    )
    headers = _auth_headers(data["usuario_id"])

    download = api_client.get(data["download_url"], headers=headers)
    assert download.status_code == 200

    review_response = api_client.post(
        (
            f"/api/v1/companies/{data['empresa_id']}/movimentos-operacionais/"
            f"lotes/{data['lote_id']}/movimentos/{data['movimentos']['aprovar']}/review"
        ),
        json={"action": "approve", "conta_final": data["contas"]["aprovacao"]},
        headers=headers,
    )
    assert review_response.status_code == 200
    assert review_response.json()["status"] == "aprovado"

    stale_feedback = _edit_feedback_sheet(
        download.content,
        {
            "stale": {
                "movimento_id": data["movimentos"]["aprovar"],
                "decisao_revisao": "corrigir",
                "contrapartida_final": data["contas"]["correcao"],
            },
        },
    )
    response = _post_feedback(api_client, data["feedback_url"], stale_feedback, headers)

    assert response.status_code == 200
    assert response.json()["total_conflitante"] == 1
    assert response.json()["total_ignorado"] == 4
    assert response.json()["total_aplicado"] == 0
    assert response.json()["resultados"][0]["status"] == "conflitante"

    with postgres_session_factory() as session:
        movimento = session.get(
            MovimentoOperacionalImportado, data["movimentos"]["aprovar"]
        )
        assert movimento.status == "aprovado"
        assert movimento.contrapartida_final == data["contas"]["aprovacao"]

    _assert_audit_metadata_safe(
        postgres_session_factory,
        empresa_id=data["empresa_id"],
        forbidden=["Historico Tratado", "DOC-TESTE"],
    )


def test_roundtrip_sem_frontend_rejeita_credenciais_legadas(
    api_client,
    postgres_session_factory,
):
    data = _seed_roundtrip_lote(
        postgres_session_factory,
        layout_version="operacional_valor_saldo_v1",
        actor_kind="legacy",
    )

    for headers in (
        {"X-API-Key": data["api_key"]},
        {"X-Admin-Token": "admin-token-legado"},
    ):
        download = api_client.get(data["download_url"], headers=headers)
        feedback = _post_feedback(
            api_client,
            data["feedback_url"],
            _edit_feedback_sheet(b"", {}),
            headers,
        )

        assert download.status_code == 401
        assert feedback.status_code == 401


def _seed_roundtrip_lote(
    postgres_session_factory, *, layout_version: str, actor_kind: str
):
    suffix = uuid4().hex[:8]
    cod_dominio = int(uuid4().int % 8_000_000) + 1_000_000
    cnpj = f"9{cod_dominio:013d}"[-14:]
    conta_financeira = int(f"10{suffix[:4]}", 16) % 80_000 + 10_000
    conta_aprovacao = int(f"20{suffix[:4]}", 16) % 80_000 + 20_000
    conta_correcao = conta_aprovacao + 1

    empresa = Empresa(
        nome_empresa=f"Empresa Roundtrip {suffix} LTDA",
        cnpj_cpf=cnpj,
        api_key=f"api-key-roundtrip-{suffix}",
        cod_dominio=cod_dominio,
    )
    usuario = Usuario(
        nome="Operador Roundtrip",
        login=f"operador.roundtrip.{actor_kind}.{suffix}",
        email=f"operador.roundtrip.{actor_kind}.{suffix}@example.com",
        senha_hash=PASSWORD_HASH.hash("senha-segura-123"),
        papel="operador",
        is_active=True,
    )
    usuario.permissoes_empresas.append(
        UsuarioEmpresaPermissao(empresa=empresa, permissao="operacao")
    )
    lote = LoteImportacaoMovimentoOperacional(
        empresa=empresa,
        usuario=usuario,
        original_filename=f"movimentos-roundtrip-{suffix}.xlsx",
        file_hash=f"sha256:roundtrip-{suffix}",
        status="completed_with_warnings",
        total_linhas=5,
        total_importadas=5,
        total_invalidas=0,
        warnings_metadata={"warnings": []},
        periodo_inicio=date(2026, 1, 1),
        periodo_fim=date(2026, 1, 31),
        cnpj_cpf_arquivo=cnpj,
        codigo_dominio_arquivo=str(cod_dominio),
        layout_version=layout_version,
    )
    movimentos = [
        _movimento(
            "aprovar",
            conta_financeira=conta_financeira,
            contrapartida_sugerida=conta_aprovacao,
            status="sugerido",
            confidence_sugerida=0.91,
        ),
        _movimento(
            "corrigir",
            conta_financeira=conta_financeira,
            status="revisao",
            mensagens_validacao=["Contrapartida pendente para revisao."],
        ),
        _movimento(
            "rejeitar",
            conta_financeira=conta_financeira,
            status="revisao",
        ),
        _movimento(
            "ignorar",
            conta_financeira=conta_financeira,
            status="revisao",
        ),
        _movimento(
            "invalida",
            conta_financeira=conta_financeira,
            status="revisao",
        ),
    ]

    with postgres_session_factory() as session:
        session.add_all(
            [
                ContaContabil(
                    codigo=conta_financeira,
                    classificacao=f"1.1.{conta_financeira}",
                    nome=f"Conta financeira {suffix}",
                    tipo="A",
                    grau=6,
                    is_active=True,
                ),
                ContaContabil(
                    codigo=conta_aprovacao,
                    classificacao=f"2.1.{conta_aprovacao}",
                    nome=f"Conta aprovacao {suffix}",
                    tipo="A",
                    grau=6,
                    is_active=True,
                ),
                ContaContabil(
                    codigo=conta_correcao,
                    classificacao=f"2.1.{conta_correcao}",
                    nome=f"Conta correcao {suffix}",
                    tipo="A",
                    grau=6,
                    is_active=True,
                ),
                lote,
            ]
        )
        session.flush()
        for linha, movimento in enumerate(movimentos, start=2):
            session.add(
                MovimentoOperacionalImportado(
                    lote_id=lote.id,
                    empresa_id=empresa.id,
                    linha_original=linha,
                    **movimento,
                )
            )
        session.commit()
        movimento_ids = {
            movimento.historico_normalizado.replace(
                "historico tratado ", ""
            ): movimento.id
            for movimento in (
                session.query(MovimentoOperacionalImportado)
                .filter_by(lote_id=lote.id)
                .all()
            )
        }
        return {
            "empresa_id": empresa.id,
            "lote_id": lote.id,
            "usuario_id": usuario.id,
            "api_key": empresa.api_key,
            "contas": {
                "aprovacao": conta_aprovacao,
                "correcao": conta_correcao,
            },
            "movimentos": movimento_ids,
            "download_url": (
                f"/api/v1/companies/{empresa.id}/movimentos-operacionais/"
                f"lotes/{lote.id}/planilha-classificada"
            ),
            "feedback_url": (
                f"/api/v1/companies/{empresa.id}/movimentos-operacionais/"
                f"lotes/{lote.id}/planilha-classificada/feedback"
            ),
        }


def _movimento(
    chave: str,
    *,
    conta_financeira: int,
    status: str,
    contrapartida_sugerida: int | None = None,
    confidence_sugerida: float | None = None,
    mensagens_validacao: list[str] | None = None,
):
    return {
        "data": date(2026, 1, 2),
        "conta_financeira": conta_financeira,
        "historico": f"Historico Tratado {chave}",
        "historico_normalizado": f"historico tratado {chave}",
        "valor_original": Decimal("-123.45"),
        "valor_absoluto": Decimal("123.45"),
        "saldo_observado_original": "1.234,56",
        "saldo_observado_decimal": Decimal("1234.56"),
        "saldo_calculado_decimal": Decimal("1234.56"),
        "warnings_saldo": [],
        "direcao": "credito",
        "tipo_movimento": "saida",
        "documento": f"DOC-TESTE-{chave}",
        "observacao": f"Observacao tratada {chave}",
        "contrapartida_informada": None,
        "contrapartida_sugerida": contrapartida_sugerida,
        "contrapartida_final": None,
        "confidence_sugerida": confidence_sugerida,
        "status": status,
        "elegivel_treino": False,
        "mensagens_validacao": mensagens_validacao or [],
        "conta_debito": None,
        "conta_credito": None,
    }


def _auth_headers(usuario_id: int) -> dict[str, str]:
    now = datetime.now(timezone.utc)
    token = jwt.encode(
        {
            "sub": str(usuario_id),
            "role": "operador",
            "type": "access",
            "iat": now,
            "exp": now + timedelta(hours=12),
        },
        settings.JWT_SECRET_KEY,
        algorithm=settings.JWT_ALGORITHM,
    )
    return {"Authorization": f"Bearer {token}"}


def _service_headers(
    postgres_session_factory, *, empresa_id: int, escopos: tuple[str, ...]
):
    with postgres_session_factory() as session:
        empresa = session.get(Empresa, empresa_id)
        identidade = IdentidadeServico(
            identifier=f"n8n-roundtrip-{empresa.cod_dominio}",
            nome="n8n Roundtrip Integracao",
            credential_hash="pendente",
            credential_fingerprint="pendente",
            status="ativa",
            empresas=[IdentidadeServicoEmpresa(empresa=empresa)],
            escopos=[
                IdentidadeServicoEscopo(escopo=escopo) for escopo in escopos
            ],
        )
        session.add(identidade)
        session.commit()
        credencial = emitir_credencial_servico(session, identidade_id=identidade.id)
        session.commit()
        return {"X-Service-Credential": credencial.secret}


def _edit_feedback_sheet(content: bytes, decisions: dict[str, dict]) -> bytes:
    if content:
        workbook = load_workbook(BytesIO(content))
    else:
        from openpyxl import Workbook

        workbook = Workbook()
        workbook.active.title = "Movimentos"
        workbook.active.append(
            [
                "lote_id",
                "movimento_id",
                "linha_original",
                "layout_version",
                "export_revision",
                "row_version",
                "decisao_revisao",
                "contrapartida_final",
                "observacao_revisao",
            ]
        )
    try:
        sheet = workbook["Movimentos"]
        headers = {cell.value: cell.column for cell in sheet[1]}
        decisions_by_id = {
            decision["movimento_id"]: decision for decision in decisions.values()
        }
        for row_index in range(2, sheet.max_row + 1):
            movimento_id = sheet.cell(row_index, headers["movimento_id"]).value
            decision = decisions_by_id.get(movimento_id)
            if not decision:
                continue
            sheet.cell(
                row_index, headers["decisao_revisao"]
            ).value = decision.get("decisao_revisao") or _decision_name(decision)
            if decision.get("contrapartida_final") is not None:
                sheet.cell(
                    row_index, headers["contrapartida_final"]
                ).value = decision["contrapartida_final"]
            sheet.cell(
                row_index, headers["observacao_revisao"]
            ).value = "Revisao via teste de integracao"

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.read()
    finally:
        workbook.close()


def _decision_name(decision: dict) -> str:
    if (
        "contrapartida_final" in decision
        and decision.get("decisao_revisao") == "corrigir"
    ):
        return "corrigir"
    if "contrapartida_final" in decision:
        return "aprovar"
    return "rejeitar"


def _post_feedback(api_client, url: str, content: bytes, headers: dict[str, str]):
    return api_client.post(
        url,
        files={"file": ("feedback.xlsx", content, XLSX_MEDIA_TYPE)},
        headers=headers,
    )


def _statuses(response) -> list[str]:
    return [item["status"] for item in response.json()["resultados"]]


def _decision_event_count(postgres_session_factory, empresa_id: int) -> int:
    with postgres_session_factory() as session:
        return (
            session.query(AuditEvent)
            .filter(
                AuditEvent.empresa_id == empresa_id,
                AuditEvent.event_type.in_(
                    [
                        "operational_movements.aprovado",
                        "operational_movements.corrigido",
                        "operational_movements.rejeitado",
                    ]
                ),
            )
            .count()
        )


def _assert_audit_metadata_safe(
    postgres_session_factory,
    *,
    empresa_id: int,
    forbidden: list[str],
) -> None:
    with postgres_session_factory() as session:
        events = session.query(AuditEvent).filter_by(empresa_id=empresa_id).all()
        assert any(
            event.event_type == "operational_movements.classified_sheet_downloaded"
            for event in events
        )
        assert any(
            event.event_type == "operational_movements.feedback_imported"
            for event in events
        )
        audit_metadata = " ".join(str(event.metadata_json) for event in events)
        for value in forbidden:
            assert value not in audit_metadata

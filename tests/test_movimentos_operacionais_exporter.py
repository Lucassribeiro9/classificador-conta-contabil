from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from core.movimentos_operacionais_exporter import (
    MovimentoOperacionalExportError,
    gerar_planilha_classificada,
)
from core.movimentos_operacionais_snapshot import (
    LoteOperacionalSnapshot,
    MovimentoOperacionalSnapshotItem,
)


def _item(
    *,
    movimento_id: int,
    linha_original: int,
    layout_version: str,
    direcao: str = "credito",
    valor_original: Decimal = Decimal("-250.75"),
    valor_absoluto: Decimal = Decimal("250.75"),
    status_atual: str = "sugerido",
    contrapartida: int | None = None,
    contrapartida_sugerida: int | None = 40101,
    contrapartida_final: int | None = None,
) -> MovimentoOperacionalSnapshotItem:
    return MovimentoOperacionalSnapshotItem(
        lote_id=10,
        movimento_id=movimento_id,
        empresa_id=1,
        linha_original=linha_original,
        layout_version=layout_version,
        export_revision="11111111-1111-4111-8111-111111111111",
        row_version=1,
        data=date(2026, 1, linha_original),
        conta_financeira=10046,
        historico=f"Historico linha {linha_original}",
        historico_normalizado=f"historico linha {linha_original}",
        valor_original=valor_original,
        valor_absoluto=valor_absoluto,
        direcao=direcao,
        tipo_movimento="saida",
        documento=f"DOC-{linha_original}",
        observacao="observacao original",
        contrapartida=contrapartida,
        contrapartida_sugerida=contrapartida_sugerida,
        confidence_sugerida=0.86,
        contrapartida_final=contrapartida_final,
        status_atual=status_atual,
        mensagem_validacao=["aguardando aprovacao"],
        saldo_observado_original=None,
        saldo_observado_decimal=None,
        saldo_calculado_decimal=None,
        warnings_saldo=[],
    )


def _snapshot(
    *,
    layout_version: str,
    movimentos: list[MovimentoOperacionalSnapshotItem],
) -> LoteOperacionalSnapshot:
    return LoteOperacionalSnapshot(
        lote_id=10,
        empresa_id=1,
        layout_version=layout_version,
        export_revision="11111111-1111-4111-8111-111111111111",
        movimentos=movimentos,
    )


def _rows_from_xlsx(content: bytes):
    workbook = load_workbook(BytesIO(content), data_only=True)
    try:
        sheet = workbook["Movimentos"]
        return list(sheet.iter_rows(values_only=True)), sheet
    finally:
        workbook.close()


def test_gera_planilha_classificada_legada_preservando_ordem_colunas_e_controles():
    layout_version = "operacional_valor_legado_v1"
    snapshot = _snapshot(
        layout_version=layout_version,
        movimentos=[
            _item(movimento_id=2, linha_original=2, layout_version=layout_version),
            _item(
                movimento_id=1,
                linha_original=1,
                layout_version=layout_version,
                status_atual="aprovado",
                contrapartida=10722,
                contrapartida_sugerida=20722,
                contrapartida_final=30722,
            ),
        ],
    )

    content = gerar_planilha_classificada(snapshot)
    rows, sheet = _rows_from_xlsx(content)

    assert sheet.freeze_panes == "A2"
    assert sheet["A1"].font.bold is True
    assert rows[0] == (
        "data",
        "conta_financeira",
        "historico",
        "valor",
        "contrapartida",
        "tipo_movimento",
        "documento",
        "observacao",
        "lote_id",
        "movimento_id",
        "linha_original",
        "layout_version",
        "export_revision",
        "row_version",
        "contrapartida_sugerida",
        "confidence_sugerida",
        "status_atual",
        "mensagem_validacao",
        "saldo_observado_original",
        "saldo_observado_decimal",
        "saldo_calculado_decimal",
        "warnings_saldo",
        "decisao_revisao",
        "contrapartida_final",
        "observacao_revisao",
    )
    assert [row[9] for row in rows[1:]] == [2, 1]
    assert rows[1][0].date() == date(2026, 1, 2)
    assert rows[1][1:8] == (
        10046,
        "Historico linha 2",
        Decimal("-250.75"),
        None,
        "saida",
        "DOC-2",
        "observacao original",
    )
    assert rows[2][4] == 10722
    assert rows[2][14:18] == (
        20722,
        0.86,
        "aprovado",
        "aguardando aprovacao",
    )
    assert rows[2][22:25] == (None, 30722, None)


def test_rejeita_layout_desconhecido():
    snapshot = _snapshot(
        layout_version="layout_desconhecido",
        movimentos=[],
    )

    with pytest.raises(MovimentoOperacionalExportError):
        gerar_planilha_classificada(snapshot)


def test_gera_layout_a_oficial_com_saldo_quando_disponivel():
    layout_version = "operacional_valor_saldo_v1"
    item = _item(
        movimento_id=1,
        linha_original=1,
        layout_version=layout_version,
        valor_original=Decimal("3660.15"),
        valor_absoluto=Decimal("3660.15"),
        direcao="debito",
    )
    item = item.__class__(
        **{
            **item.__dict__,
            "saldo_observado_original": "10.000,00",
            "saldo_observado_decimal": Decimal("10000.00"),
            "saldo_calculado_decimal": Decimal("10000.00"),
        }
    )

    content = gerar_planilha_classificada(
        _snapshot(layout_version=layout_version, movimentos=[item])
    )
    rows, _ = _rows_from_xlsx(content)

    assert rows[0][:9] == (
        "data",
        "conta_financeira",
        "historico",
        "valor",
        "saldo",
        "contrapartida",
        "tipo_movimento",
        "documento",
        "observacao",
    )
    assert rows[1][3] == pytest.approx(3660.15)
    assert rows[1][4:6] == ("10.000,00", None)


def test_gera_layout_b_oficial_separando_debito_e_credito_por_direcao():
    layout_version = "operacional_debito_credito_saldo_v1"
    snapshot = _snapshot(
        layout_version=layout_version,
        movimentos=[
            _item(
                movimento_id=1,
                linha_original=1,
                layout_version=layout_version,
                direcao="credito",
                valor_original=Decimal("-250.75"),
                valor_absoluto=Decimal("250.75"),
            ),
            _item(
                movimento_id=2,
                linha_original=2,
                layout_version=layout_version,
                direcao="debito",
                valor_original=Decimal("3660.15"),
                valor_absoluto=Decimal("3660.15"),
            ),
        ],
    )

    content = gerar_planilha_classificada(snapshot)
    rows, _ = _rows_from_xlsx(content)

    assert rows[0][:10] == (
        "data",
        "conta_financeira",
        "historico",
        "debito",
        "credito",
        "saldo",
        "contrapartida",
        "tipo_movimento",
        "documento",
        "observacao",
    )
    assert rows[1][3] == pytest.approx(250.75)
    assert rows[1][4:6] == (None, None)
    assert rows[2][3] is None
    assert rows[2][4] == pytest.approx(3660.15)
    assert rows[2][5] is None


def test_planilha_protege_colunas_somente_leitura_e_deixa_editaveis_destravadas():
    layout_version = "operacional_valor_legado_v1"
    content = gerar_planilha_classificada(
        _snapshot(
            layout_version=layout_version,
            movimentos=[
                _item(movimento_id=1, linha_original=1, layout_version=layout_version)
            ],
        )
    )
    workbook = load_workbook(BytesIO(content), data_only=True)
    try:
        sheet = workbook["Movimentos"]
        headers = [cell.value for cell in sheet[1]]
        decisao_col = headers.index("decisao_revisao") + 1
        contrapartida_final_col = headers.index("contrapartida_final") + 1
        status_col = headers.index("status_atual") + 1
        movimento_id_col = headers.index("movimento_id") + 1

        assert sheet.protection.sheet is True
        assert sheet.cell(row=2, column=decisao_col).protection.locked is False
        assert (
            sheet.cell(row=2, column=contrapartida_final_col).protection.locked
            is False
        )
        assert sheet.cell(row=2, column=status_col).protection.locked is True
        assert sheet.cell(row=2, column=movimento_id_col).protection.locked is True
    finally:
        workbook.close()


def test_planilha_exporta_apenas_linhas_presentes_no_snapshot_recebido():
    layout_version = "operacional_valor_legado_v1"
    content = gerar_planilha_classificada(
        _snapshot(
            layout_version=layout_version,
            movimentos=[
                _item(movimento_id=1, linha_original=1, layout_version=layout_version)
            ],
        )
    )

    rows, _ = _rows_from_xlsx(content)
    headers = rows[0]
    empresa_id_col = headers.index("empresa_id") if "empresa_id" in headers else None

    assert len(rows) == 2
    assert rows[1][headers.index("movimento_id")] == 1
    assert empresa_id_col is None

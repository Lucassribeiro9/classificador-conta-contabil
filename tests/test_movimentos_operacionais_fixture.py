from pathlib import Path

from openpyxl import load_workbook


LEGACY_FIXTURE_PATH = (
    Path(__file__).parent
    / "fixtures"
    / "modelo_movimentos_operacionais_classificacao.xlsx"
)
VALOR_SALDO_FIXTURE_PATH = (
    Path(__file__).parent
    / "fixtures"
    / "modelo_movimentos_operacionais_valor_saldo.xlsx"
)
DEBITO_CREDITO_SALDO_FIXTURE_PATH = (
    Path(__file__).parent
    / "fixtures"
    / "modelo_movimentos_operacionais_debito_credito_saldo.xlsx"
)

EXPECTED_SHEETS = ["Movimentos", "Instrucoes", "Exemplos"]
METADATA_LABELS = {
    "Empresa",
    "Codigo dominio",
    "CNPJ/CPF",
    "Periodo inicio",
    "Periodo fim",
}
LEGACY_HEADERS = [
    "data",
    "conta_financeira",
    "historico",
    "valor",
    "contrapartida",
    "tipo_movimento",
    "documento",
    "observacao",
    "status_sugerido",
    "confidence_sugerida",
    "mensagem_validacao",
]
VALOR_SALDO_HEADERS = [
    "data",
    "conta_financeira",
    "historico",
    "valor",
    "saldo",
    "contrapartida",
    "tipo_movimento",
    "documento",
    "observacao",
    "status_sugerido",
    "confidence_sugerida",
    "mensagem_validacao",
]
DEBITO_CREDITO_SALDO_HEADERS = [
    "data",
    "conta_financeira",
    "historico",
    "debito",
    "credito",
    "saldo",
    "contrapartida",
    "tipo_movimento",
    "documento",
    "observacao",
    "status_sugerido",
    "confidence_sugerida",
    "mensagem_validacao",
]

MODEL_FIXTURES = [
    (LEGACY_FIXTURE_PATH, LEGACY_HEADERS),
    (VALOR_SALDO_FIXTURE_PATH, VALOR_SALDO_HEADERS),
    (DEBITO_CREDITO_SALDO_FIXTURE_PATH, DEBITO_CREDITO_SALDO_HEADERS),
]


def _find_header_row(sheet):
    required = {"data", "conta_financeira", "historico"}
    for row in sheet.iter_rows(values_only=True):
        values = [str(cell).strip() if cell is not None else "" for cell in row]
        if required.issubset(values):
            return values
    return None


def _metadata_labels(sheet):
    return {
        str(row[0]).strip()
        for row in sheet.iter_rows(min_row=1, max_row=8, values_only=True)
        if row and row[0] is not None
    }


def test_modelos_movimentos_operacionais_estao_versionados():
    for fixture_path, _headers in MODEL_FIXTURES:
        assert fixture_path.exists(), f"Fixture ausente: {fixture_path.name}"


def test_modelos_movimentos_operacionais_definem_contratos_de_headers():
    for fixture_path, expected_headers in MODEL_FIXTURES:
        workbook = load_workbook(fixture_path, read_only=True, data_only=True)
        try:
            assert workbook.sheetnames == EXPECTED_SHEETS

            sheet = workbook["Movimentos"]
            assert METADATA_LABELS.issubset(_metadata_labels(sheet))

            headers = _find_header_row(sheet)
            assert headers == expected_headers
        finally:
            workbook.close()


def test_modelos_movimentos_operacionais_nao_trazem_dados_reais():
    forbidden_fragments = {
        "empresa brasileira de beneficios",
        "pagto instituicao",
        "lucas",
        "ribeiro",
        "dominio contabilidade",
    }

    for fixture_path, _headers in MODEL_FIXTURES:
        workbook = load_workbook(fixture_path, read_only=True, data_only=True)
        try:
            sheet = workbook["Movimentos"]

            assert sheet["B2"].value is None
            assert sheet["B3"].value is None
            assert sheet["B4"].value is None
            assert sheet["B5"].value is None
            assert sheet["B6"].value is None

            for row in sheet.iter_rows(values_only=True):
                for value in row:
                    if value is None:
                        continue
                    normalized = str(value).strip().lower()
                    assert all(
                        fragment not in normalized for fragment in forbidden_fragments
                    )
        finally:
            workbook.close()


def test_modelos_movimentos_operacionais_trazem_exemplos_ficticios():
    for fixture_path in [VALOR_SALDO_FIXTURE_PATH, DEBITO_CREDITO_SALDO_FIXTURE_PATH]:
        workbook = load_workbook(fixture_path, read_only=True, data_only=True)
        try:
            sheet = workbook["Exemplos"]

            rows = list(sheet.iter_rows(values_only=True))
            assert any(
                "exemplo" in str(cell).lower()
                for row in rows
                for cell in row
                if cell
            )
            assert any(
                "saldo" in str(cell).lower()
                for row in rows
                for cell in row
                if cell
            )
        finally:
            workbook.close()


def _sheet_text(sheet):
    return " ".join(
        str(cell).strip().lower()
        for row in sheet.iter_rows(values_only=True)
        for cell in row
        if cell is not None
    )


def test_modelos_oficiais_instrucoes_e_exemplos_usam_apenas_dados_ficticios():
    forbidden_fragments = {
        "empresa brasileira de beneficios",
        "pagto instituicao",
        "lucas",
        "ribeiro",
        "dominio contabilidade",
        "cnpj real",
        "cpf real",
        "token real",
    }

    for fixture_path in [VALOR_SALDO_FIXTURE_PATH, DEBITO_CREDITO_SALDO_FIXTURE_PATH]:
        workbook = load_workbook(fixture_path, read_only=True, data_only=True)
        try:
            for sheet_name in ["Instrucoes", "Exemplos"]:
                content = _sheet_text(workbook[sheet_name])

                for fragment in forbidden_fragments:
                    assert fragment not in content
        finally:
            workbook.close()


def test_modelos_oficiais_trazem_instrucoes_semanticas_equivalentes():
    expected_terms = {
        "objetivo",
        "conta financeira",
        "contrapartida",
        "saldo",
        "derivacao",
        "validacao",
        "revisao",
        "status_sugerido",
        "confidence_sugerida",
        "mensagem_validacao",
    }

    for fixture_path in [VALOR_SALDO_FIXTURE_PATH, DEBITO_CREDITO_SALDO_FIXTURE_PATH]:
        workbook = load_workbook(fixture_path, read_only=True, data_only=True)
        try:
            instructions = _sheet_text(workbook["Instrucoes"])

            for term in expected_terms:
                assert term in instructions
        finally:
            workbook.close()


def test_modelo_valor_saldo_explica_convencao_de_entrada_e_saida():
    workbook = load_workbook(VALOR_SALDO_FIXTURE_PATH, read_only=True, data_only=True)
    try:
        instructions = _sheet_text(workbook["Instrucoes"])

        assert "valor positivo" in instructions
        assert "entrada" in instructions
        assert "valor negativo" in instructions
        assert "saida" in instructions
    finally:
        workbook.close()


def test_modelo_debito_credito_saldo_explica_convencao_de_extrato():
    workbook = load_workbook(
        DEBITO_CREDITO_SALDO_FIXTURE_PATH, read_only=True, data_only=True
    )
    try:
        instructions = _sheet_text(workbook["Instrucoes"])

        assert "credito" in instructions
        assert "entrada" in instructions
        assert "debito" in instructions
        assert "saida" in instructions
        assert "exatamente uma" in instructions
    finally:
        workbook.close()


def _example_rows(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[1]]
    return [
        dict(zip(headers, row))
        for row in rows[2:]
        if row and any(cell is not None for cell in row)
    ]


def test_modelos_oficiais_exemplos_cobrem_cenarios_operacionais():
    expected_scenarios = {
        "recebimento",
        "pagamento",
        "aplicacao",
        "resgate",
        "classificar",
    }

    for fixture_path in [VALOR_SALDO_FIXTURE_PATH, DEBITO_CREDITO_SALDO_FIXTURE_PATH]:
        workbook = load_workbook(fixture_path, read_only=True, data_only=True)
        try:
            scenarios = {
                str(row["cenario"]).strip().lower()
                for row in _example_rows(workbook["Exemplos"])
            }

            assert expected_scenarios == scenarios
        finally:
            workbook.close()


def _fill_color(cell):
    rgb = cell.fill.fgColor.rgb or ""
    return rgb[-6:]


def test_modelos_oficiais_movimentos_trazem_faixa_de_orientacao():
    for fixture_path in [VALOR_SALDO_FIXTURE_PATH, DEBITO_CREDITO_SALDO_FIXTURE_PATH]:
        workbook = load_workbook(fixture_path, read_only=True, data_only=True)
        try:
            sheet = workbook["Movimentos"]
            orientation = _sheet_text(sheet)

            assert "preencha uma linha por movimento" in orientation
            assert "colunas amarelas sao obrigatorias" in orientation
            assert "verdes sao opcionais" in orientation
            assert "azuis podem ser preenchidas pelo sistema" in orientation
        finally:
            workbook.close()


def test_modelos_oficiais_movimentos_trazem_orientacao_por_coluna():
    required_descriptions = {
        "data do movimento",
        "conta financeira",
        "historico original",
        "saldo observado",
        "contrapartida",
        "preenchido pelo sistema",
    }

    for fixture_path in [VALOR_SALDO_FIXTURE_PATH, DEBITO_CREDITO_SALDO_FIXTURE_PATH]:
        workbook = load_workbook(fixture_path, read_only=True, data_only=True)
        try:
            sheet = workbook["Movimentos"]
            header = _find_header_row(sheet)
            header_index = next(
                row_number
                for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1)
                if [str(cell).strip() if cell is not None else "" for cell in row]
                == header
            )
            descriptions = " ".join(
                str(cell.value).strip().lower()
                for cell in sheet[header_index + 1]
                if cell.value is not None
            )

            for description in required_descriptions:
                assert description in descriptions
        finally:
            workbook.close()


def test_modelos_oficiais_movimentos_colorem_colunas_por_tipo_de_preenchimento():
    expected_colors = {
        VALOR_SALDO_FIXTURE_PATH: {
            "data": "FFF2CC",
            "conta_financeira": "FFF2CC",
            "historico": "FFF2CC",
            "valor": "FFF2CC",
            "saldo": "E2F0D9",
            "contrapartida": "E2F0D9",
            "tipo_movimento": "E2F0D9",
            "documento": "E2F0D9",
            "observacao": "E2F0D9",
            "status_sugerido": "D9EAF7",
            "confidence_sugerida": "D9EAF7",
            "mensagem_validacao": "D9EAF7",
        },
        DEBITO_CREDITO_SALDO_FIXTURE_PATH: {
            "data": "FFF2CC",
            "conta_financeira": "FFF2CC",
            "historico": "FFF2CC",
            "debito": "FFF2CC",
            "credito": "FFF2CC",
            "saldo": "E2F0D9",
            "contrapartida": "E2F0D9",
            "tipo_movimento": "E2F0D9",
            "documento": "E2F0D9",
            "observacao": "E2F0D9",
            "status_sugerido": "D9EAF7",
            "confidence_sugerida": "D9EAF7",
            "mensagem_validacao": "D9EAF7",
        },
    }

    for fixture_path, colors_by_header in expected_colors.items():
        workbook = load_workbook(fixture_path, read_only=True, data_only=True)
        try:
            sheet = workbook["Movimentos"]
            headers = _find_header_row(sheet)
            header_index = next(
                row_number
                for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1)
                if [str(cell).strip() if cell is not None else "" for cell in row]
                == headers
            )

            for column_index, header in enumerate(headers, 1):
                assert (
                    _fill_color(sheet.cell(header_index, column_index))
                    == colors_by_header[header]
                )
        finally:
            workbook.close()


def test_modelos_oficiais_exemplos_mantem_movimentos_vazia():
    for fixture_path in [VALOR_SALDO_FIXTURE_PATH, DEBITO_CREDITO_SALDO_FIXTURE_PATH]:
        workbook = load_workbook(fixture_path, read_only=True, data_only=True)
        try:
            sheet = workbook["Movimentos"]
            header_row = _find_header_row(sheet)
            header_index = next(
                row_number
                for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1)
                if [str(cell).strip() if cell is not None else "" for cell in row]
                == header_row
            )
            rows_after_orientation = list(sheet.iter_rows(min_row=header_index + 2))

            assert rows_after_orientation == []
        finally:
            workbook.close()


def test_modelos_oficiais_exemplos_tem_saldos_e_derivacoes_coerentes():
    valor_workbook = load_workbook(
        VALOR_SALDO_FIXTURE_PATH, read_only=True, data_only=True
    )
    debito_credito_workbook = load_workbook(
        DEBITO_CREDITO_SALDO_FIXTURE_PATH, read_only=True, data_only=True
    )
    try:
        valor_rows = _example_rows(valor_workbook["Exemplos"])
        debito_credito_rows = _example_rows(debito_credito_workbook["Exemplos"])

        assert len(valor_rows) == len(debito_credito_rows)

        previous_balance = None
        for valor_row, debito_credito_row in zip(valor_rows, debito_credito_rows):
            assert valor_row["cenario"] == debito_credito_row["cenario"]
            assert valor_row["conta_financeira"] == debito_credito_row["conta_financeira"]
            assert valor_row["historico"] == debito_credito_row["historico"]
            assert valor_row["contrapartida"] == debito_credito_row["contrapartida"]
            assert valor_row["debito_derivado"] == debito_credito_row["debito_derivado"]
            assert valor_row["credito_derivado"] == debito_credito_row["credito_derivado"]

            valor = valor_row["valor"]
            debito = debito_credito_row["debito"] or 0
            credito = debito_credito_row["credito"] or 0
            assert (debito == 0) != (credito == 0)
            assert credito - debito == valor

            if previous_balance is not None:
                assert valor_row["saldo"] == previous_balance + valor
                assert debito_credito_row["saldo"] == previous_balance + valor
            assert valor_row["saldo"] == debito_credito_row["saldo"]
            previous_balance = valor_row["saldo"]
    finally:
        valor_workbook.close()
        debito_credito_workbook.close()


def test_modelo_legado_movimentos_operacionais_permanece_sem_saldo():
    workbook = load_workbook(LEGACY_FIXTURE_PATH, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == EXPECTED_SHEETS

        sheet = workbook["Movimentos"]
        assert METADATA_LABELS.issubset(_metadata_labels(sheet))

        headers = _find_header_row(sheet)
        assert headers == LEGACY_HEADERS
        assert "saldo" not in headers

    finally:
        workbook.close()

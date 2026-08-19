from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl import load_workbook

from core.movimentos_operacionais_parser import (
    MovimentoOperacionalParseError,
    parse_movimentos_operacionais_xlsx,
)


FIXTURES_DIR = Path(__file__).parent / "fixtures"
MODELO_PATH = FIXTURES_DIR / "modelo_movimentos_operacionais_classificacao.xlsx"
MODELO_VALOR_SALDO_PATH = (
    FIXTURES_DIR / "modelo_movimentos_operacionais_valor_saldo.xlsx"
)
MODELO_DEBITO_CREDITO_SALDO_PATH = (
    FIXTURES_DIR / "modelo_movimentos_operacionais_debito_credito_saldo.xlsx"
)


def _modelo_preenchido(tmp_path):
    """Cria uma copia temporaria do modelo com metadados ficticios."""

    path = tmp_path / "movimentos-operacionais-validos.xlsx"
    workbook = load_workbook(MODELO_PATH)
    sheet = workbook["Movimentos"]
    sheet["B2"] = "EMPRESA OPERACIONAL TESTE LTDA"
    sheet["B3"] = "1122"
    sheet["B4"] = "11.222.333/0001-44"
    sheet["B5"] = "01/01/2025"
    sheet["B6"] = "31/01/2025"
    workbook.save(path)
    workbook.close()
    return path


def _modelo_oficial_preenchido(tmp_path, modelo_path):
    """Cria uma copia temporaria de um modelo oficial com metadados ficticios."""

    path = tmp_path / modelo_path.name
    workbook = load_workbook(modelo_path)
    sheet = workbook["Movimentos"]
    sheet["B2"] = "EMPRESA OPERACIONAL TESTE LTDA"
    sheet["B3"] = "1122"
    sheet["B4"] = "11.222.333/0001-44"
    sheet["B5"] = "01/01/2025"
    sheet["B6"] = "31/01/2025"
    workbook.save(path)
    workbook.close()
    return path


def _write_workbook(path, rows, *, sheet_name="Movimentos"):
    """Grava uma planilha simples para cenarios de layout invalido."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()


def test_parse_movimentos_operacionais_detecta_layout_valor_saldo(tmp_path):
    """Deve identificar layout oficial com valor assinado e saldo."""

    xlsx_path = _modelo_oficial_preenchido(tmp_path, MODELO_VALOR_SALDO_PATH)

    result = parse_movimentos_operacionais_xlsx(xlsx_path)

    assert result.layout_version == "operacional_valor_saldo_v1"


def test_parse_movimentos_operacionais_detecta_layout_debito_credito_saldo(tmp_path):
    """Deve identificar layout oficial com debito, credito e saldo."""

    xlsx_path = _modelo_oficial_preenchido(
        tmp_path,
        MODELO_DEBITO_CREDITO_SALDO_PATH,
    )

    result = parse_movimentos_operacionais_xlsx(xlsx_path)

    assert result.layout_version == "operacional_debito_credito_saldo_v1"


def test_parse_movimentos_operacionais_detecta_layout_legado(tmp_path):
    """Deve identificar o modelo legado sem saldo como compatibilidade."""

    xlsx_path = _modelo_preenchido(tmp_path)

    result = parse_movimentos_operacionais_xlsx(xlsx_path)

    assert result.layout_version == "operacional_valor_legado_v1"


def test_parse_movimentos_operacionais_rejeita_layout_ambiguo(tmp_path):
    """Deve rejeitar cabecalho com valor e debito/credito no mesmo layout."""

    xlsx_path = tmp_path / "layout-ambiguo.xlsx"
    _write_workbook(
        xlsx_path,
        [
            ["Empresa", "EMPRESA OPERACIONAL TESTE LTDA"],
            ["Codigo dominio", "1122"],
            ["CNPJ/CPF", "11.222.333/0001-44"],
            ["Periodo inicio", "01/01/2025"],
            ["Periodo fim", "31/01/2025"],
            [],
            [
                "data",
                "conta_financeira",
                "historico",
                "valor",
                "debito",
                "credito",
            ],
        ],
    )

    with pytest.raises(
        MovimentoOperacionalParseError,
        match="valor.*debito.*credito",
    ):
        parse_movimentos_operacionais_xlsx(xlsx_path)


def test_parse_movimentos_operacionais_normaliza_credito_do_layout_b(tmp_path):
    """Deve converter credito do extrato em valor interno positivo."""

    xlsx_path = tmp_path / "layout-b-credito.xlsx"
    _write_workbook(
        xlsx_path,
        [
            ["Empresa", "EMPRESA OPERACIONAL TESTE LTDA"],
            ["Codigo dominio", "1122"],
            ["CNPJ/CPF", "11.222.333/0001-44"],
            ["Periodo inicio", "01/01/2025"],
            ["Periodo fim", "31/01/2025"],
            [],
            [
                "data",
                "conta_financeira",
                "historico",
                "debito",
                "credito",
                "saldo",
            ],
            ["02/01/2025", 10046, "RECEBIMENTO CLIENTE", None, 1500, 2500],
        ],
    )

    result = parse_movimentos_operacionais_xlsx(xlsx_path)

    assert result.movimentos[0]["valor"] == 1500
    assert "debito" not in result.movimentos[0]
    assert "credito" not in result.movimentos[0]


def test_parse_movimentos_operacionais_normaliza_debito_do_layout_b(tmp_path):
    """Deve converter debito do extrato em valor interno negativo."""

    xlsx_path = tmp_path / "layout-b-debito.xlsx"
    _write_workbook(
        xlsx_path,
        [
            ["Empresa", "EMPRESA OPERACIONAL TESTE LTDA"],
            ["Codigo dominio", "1122"],
            ["CNPJ/CPF", "11.222.333/0001-44"],
            ["Periodo inicio", "01/01/2025"],
            ["Periodo fim", "31/01/2025"],
            [],
            [
                "data",
                "conta_financeira",
                "historico",
                "debito",
                "credito",
                "saldo",
            ],
            ["03/01/2025", 10046, "PAGAMENTO FORNECEDOR", 300, None, 2200],
        ],
    )

    result = parse_movimentos_operacionais_xlsx(xlsx_path)

    assert result.movimentos[0]["valor"] == -300
    assert "debito" not in result.movimentos[0]
    assert "credito" not in result.movimentos[0]


def test_parse_movimentos_operacionais_ler_fixture_preenchida(tmp_path):
    """Deve ler a fixture preenchida e retornar metadados e movimentos."""

    xlsx_path = _modelo_preenchido(tmp_path)

    result = parse_movimentos_operacionais_xlsx(xlsx_path)

    assert result.metadata.empresa_nome == "EMPRESA OPERACIONAL TESTE LTDA"
    assert result.metadata.codigo_dominio == "1122"
    assert result.metadata.cnpj_cpf == "11222333000144"
    assert result.metadata.periodo_inicio == "2025-01-01"
    assert result.metadata.periodo_fim == "2025-01-31"
    assert result.movimentos[0] == {
        "data": "2025-01-02",
        "conta_financeira": 10046,
        "historico": "RECEBTO.DUPLICATAS",
        "valor": 3660.15,
        "contrapartida": 10722,
        "tipo_movimento": "entrada",
        "documento": "OFX-0001",
        "observacao": "Contrapartida conhecida pelo contador",
    }
    assert result.movimentos[2]["contrapartida"] is None
    assert len(result.movimentos) == 5


def test_parse_movimentos_operacionais_rejeita_arquivo_sem_aba_movimentos(tmp_path):
    """Deve rejeitar arquivo sem a aba obrigatoria Movimentos."""

    xlsx_path = tmp_path / "sem-aba-movimentos.xlsx"
    _write_workbook(xlsx_path, [["CNPJ/CPF", "11.222.333/0001-44"]], sheet_name="Dados")

    with pytest.raises(
        MovimentoOperacionalParseError,
        match="aba Movimentos nao encontrada",
    ):
        parse_movimentos_operacionais_xlsx(xlsx_path)


def test_parse_movimentos_operacionais_rejeita_lote_sem_cnpj():
    """Deve rejeitar lote sem CNPJ/CPF preenchido."""

    with pytest.raises(
        MovimentoOperacionalParseError,
        match="metadados obrigatorios: cnpj_cpf",
    ):
        parse_movimentos_operacionais_xlsx(MODELO_PATH)


def test_parse_movimentos_operacionais_rejeita_lote_sem_periodo_valido(tmp_path):
    """Deve rejeitar lote quando o periodo obrigatorio for invalido."""

    xlsx_path = _modelo_preenchido(tmp_path)
    workbook = load_workbook(xlsx_path)
    sheet = workbook["Movimentos"]
    sheet["B5"] = "31/02/2025"
    workbook.save(xlsx_path)
    workbook.close()

    with pytest.raises(
        MovimentoOperacionalParseError,
        match="metadados obrigatorios: periodo_inicio",
    ):
        parse_movimentos_operacionais_xlsx(xlsx_path)


def test_parse_movimentos_operacionais_rejeita_cabecalho_sem_coluna_obrigatoria(
    tmp_path,
):
    """Deve rejeitar cabecalho que nao contenha todas as colunas obrigatorias."""

    xlsx_path = tmp_path / "sem-coluna-obrigatoria.xlsx"
    _write_workbook(
        xlsx_path,
        [
            ["Empresa", "EMPRESA OPERACIONAL TESTE LTDA"],
            ["Codigo dominio", "1122"],
            ["CNPJ/CPF", "11.222.333/0001-44"],
            ["Periodo inicio", "01/01/2025"],
            ["Periodo fim", "31/01/2025"],
            [],
            ["data", "conta_financeira", "historico", "contrapartida"],
            ["02/01/2025", 10046, "RECEBTO.DUPLICATAS", 10722],
        ],
    )

    with pytest.raises(
        MovimentoOperacionalParseError,
        match="cabecalho com colunas obrigatorias nao encontrado",
    ):
        parse_movimentos_operacionais_xlsx(xlsx_path)

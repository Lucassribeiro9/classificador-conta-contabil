import re
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from core.database import Base
from core.dataset_builder import build_dataset_treino_contrapartida
from core.ml_engine import ClassificadorContabil
from core.models import Empresa, MovimentoOperacionalImportado, Usuario
from core.movimentos_operacionais_classification import (
    classificar_movimentos_operacionais_pendentes,
)
from core.movimentos_operacionais_importer import import_movimentos_operacionais
from core.movimentos_operacionais_parser import parse_movimentos_operacionais_xlsx
from core.plano_contas_financeiro import infer_is_financial_origin
from core.plano_contas_importer import import_plano_contas
from core.plano_contas_parser import parse_plano_contas_xlsx
from core.razao_importer import import_razao
from core.razao_parser import parse_razao_xlsx_with_metadata


FIXTURES_DIR = Path(__file__).parent / "fixtures" / "homologacao"
SANITIZED_CNPJ = "22333444000155"
EXPECTED_FILES = {
    "README.md",
    "movimentos_operacionais_hml.xlsx",
    "plano_contas_hml.xlsx",
    "razao_hml.xlsx",
}


@pytest.fixture()
def session():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(bind=engine)
    testing_session = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    db = testing_session()
    try:
        yield db
    finally:
        db.close()
        Base.metadata.drop_all(bind=engine)
        engine.dispose()


def test_homologacao_fixture_set_is_versioned():
    existing_files = (
        {path.name for path in FIXTURES_DIR.iterdir()}
        if FIXTURES_DIR.exists()
        else set()
    )

    assert EXPECTED_FILES <= existing_files


def test_homologacao_fixtures_are_parseable_and_coherent():
    contas = parse_plano_contas_xlsx(FIXTURES_DIR / "plano_contas_hml.xlsx")
    razao = parse_razao_xlsx_with_metadata(FIXTURES_DIR / "razao_hml.xlsx")
    movimentos = parse_movimentos_operacionais_xlsx(
        FIXTURES_DIR / "movimentos_operacionais_hml.xlsx"
    )

    codigos = {conta["codigo"] for conta in contas}
    origens_financeiras = {
        conta["codigo"]
        for conta in contas
        if infer_is_financial_origin(conta["nome"], conta["classificacao"])
    }

    assert {conta["tipo"] for conta in contas} == {"A", "S"}
    assert len(origens_financeiras) >= 3
    assert razao.metadata.cnpj_cpf == SANITIZED_CNPJ
    assert movimentos.metadata.cnpj_cpf == SANITIZED_CNPJ
    assert movimentos.metadata.codigo_dominio == "7701"
    assert {lancamento["conta_origem"] for lancamento in razao.lancamentos} <= {
        str(codigo) for codigo in origens_financeiras
    }
    assert {
        lancamento["contrapartida"] for lancamento in razao.lancamentos
    } <= {str(codigo) for codigo in codigos}
    assert any(lancamento["debito"] for lancamento in razao.lancamentos)
    assert any(lancamento["credito"] for lancamento in razao.lancamentos)
    assert {movimento["conta_financeira"] for movimento in movimentos.movimentos} <= (
        origens_financeiras
    )
    assert {
        movimento["contrapartida"]
        for movimento in movimentos.movimentos
        if movimento["contrapartida"] is not None
    } <= codigos


def test_homologacao_fixtures_complete_importer_flow(session, tmp_path):
    plano = parse_plano_contas_xlsx(FIXTURES_DIR / "plano_contas_hml.xlsx")
    plano_result = import_plano_contas(session, plano)
    empresa = Empresa(
        nome_empresa="EMPRESA MODELO HOMOLOGACAO LTDA",
        cnpj_cpf=SANITIZED_CNPJ,
        api_key="api-key-ficticia-hml",
        cod_dominio=7701,
    )
    usuario = Usuario(
        nome="OPERADOR MODELO HOMOLOGACAO",
        login="operador.hml",
        email="operador.hml@example.invalid",
        senha_hash="$argon2id$v=19$hash-ficticio-hml",
        papel="operador",
    )
    session.add_all([empresa, usuario])
    session.flush()

    razao_result = import_razao(
        session,
        FIXTURES_DIR / "razao_hml.xlsx",
        empresa_id=empresa.id,
        usuario_id=usuario.id,
        original_filename="razao_hml.xlsx",
    )
    movimentos_result = import_movimentos_operacionais(
        session,
        FIXTURES_DIR / "movimentos_operacionais_hml.xlsx",
        empresa_id=empresa.id,
        usuario_id=usuario.id,
        original_filename="movimentos_operacionais_hml.xlsx",
    )
    movimentos = (
        session.query(MovimentoOperacionalImportado)
        .order_by(MovimentoOperacionalImportado.documento)
        .all()
    )

    assert plano_result.criadas == len(plano)
    assert plano_result.invalidas == 0
    assert razao_result.status == "completed"
    assert razao_result.total_importadas == 12
    assert movimentos_result.status == "completed_with_warnings"
    assert movimentos_result.total_importadas == 5
    assert movimentos_result.total_invalidas == 0
    assert [movimento.status for movimento in movimentos] == [
        "pre_classificado",
        "pre_classificado",
        "pendente",
        "revisao",
        "pre_classificado",
    ]
    assert movimentos_result.warnings == [
        {
            "linha": 1,
            "warnings": [
                "Saldo ausente; conferencia por saldo limitada para esta linha.",
                "Saldo inicial ausente; saldo calculado partiu de zero.",
            ],
        },
        {
            "linha": 2,
            "warnings": [
                "Saldo ausente; conferencia por saldo limitada para esta linha."
            ],
        },
        {
            "linha": 3,
            "warnings": [
                "Saldo ausente; conferencia por saldo limitada para esta linha."
            ],
        },
        {
            "linha": 4,
            "warnings": [
                "Tipo de movimento transferencia exige contrapartida."
            ],
        },
        {
            "linha": 4,
            "warnings": [
                "Saldo ausente; conferencia por saldo limitada para esta linha."
            ],
        },
        {
            "linha": 5,
            "warnings": [
                "Saldo ausente; conferencia por saldo limitada para esta linha.",
                "Saldo inicial ausente; saldo calculado partiu de zero.",
            ],
        },
    ]

    dataset = build_dataset_treino_contrapartida(session, empresa_id=empresa.id)
    classifier = ClassificadorContabil(session, model_dir=tmp_path)
    assert classifier.train_from_dataset(dataset) is True

    classificacao = classificar_movimentos_operacionais_pendentes(
        session,
        empresa_id=empresa.id,
        model_dir=tmp_path,
    )
    movimento_sugerido = next(
        movimento for movimento in movimentos if movimento.documento == "HML-003"
    )

    assert classificacao["total_sugerido"] == 1
    assert classificacao["total_revisao"] == 0
    assert movimento_sugerido.contrapartida_sugerida == 40101
    assert movimento_sugerido.confidence_sugerida >= 0.70


def test_homologacao_fixtures_keep_only_sanitized_static_data():
    identifiers = set()

    for fixture_path in FIXTURES_DIR.glob("*.xlsx"):
        with ZipFile(fixture_path) as archive:
            assert not any(
                name.startswith("xl/externalLinks/") for name in archive.namelist()
            )

        workbook = load_workbook(fixture_path, read_only=False, data_only=False)
        try:
            for sheet in workbook.worksheets:
                assert sheet.sheet_state == "visible"
                for row in sheet.iter_rows():
                    for cell in row:
                        value = cell.value
                        if not isinstance(value, str):
                            continue
                        assert not value.startswith("=")
                        digits = re.sub(r"\D", "", value)
                        if len(digits) == 14:
                            identifiers.add(digits)
        finally:
            workbook.close()

    assert identifiers == {SANITIZED_CNPJ}
